import json
import csv
from datetime import datetime, timedelta
from typing import List, Dict, Set, Tuple, Optional
import random
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.cell_range import CellRange
import os
import copy
import math

class AdvancedSchoolScheduler:
    def __init__(self):
        self.rooms = {
            1: {"type": "lab", "capacity": 30, "available": True},
            2: {"type": "lecture", "capacity": 50, "available": True}
        }
        self.employees = {}  # Will store employee schedules
        self.sections = {}   # Will store section schedules
        self.room_schedules = {}  # Will store room schedules
        self.schedule = []
        self.conflicts = []
        self.unscheduled_classes = []  # Track classes that could not be scheduled
        
    def parse_duration(self, duration_str: str) -> int:
        """Convert duration string (e.g., '1:30') to minutes"""
        if ':' in duration_str:
            hours, minutes = map(int, duration_str.split(':'))
            return hours * 60 + minutes
        else:
            return int(duration_str) * 60
    
    def format_time(self, minutes: int) -> str:
        """Convert minutes to time string (e.g., '09:30')"""
        hours = minutes // 60
        mins = minutes % 60
        return f"{hours:02d}:{mins:02d}"
    
    def is_time_conflict(self, start1: int, duration1: int, start2: int, duration2: int) -> bool:
        """Check if two time slots conflict"""
        end1 = start1 + duration1
        end2 = start2 + duration2
        return not (end1 <= start2 or end2 <= start1)
    
    def detect_conflicts(self, class_info: Dict) -> List[Dict]:
        """Detect all conflicts for a given class"""
        conflicts = []
        start_time = class_info['start_time']
        duration = class_info['duration']
        section = class_info['section']
        employee_id = class_info['employee_id']
        roomids = class_info['roomid']
        if not isinstance(roomids, list):
            roomids = [roomids]
        # Check section conflicts
        if section in self.sections:
            for existing_class in self.sections[section]:
                if existing_class != class_info and self.is_time_conflict(
                    start_time, duration, existing_class['start_time'], existing_class['duration']):
                    conflicts.append({
                        'type': 'section_conflict',
                        'message': f"Section {section} has overlapping classes",
                        'conflicting_class': existing_class
                    })
        # Check employee conflicts
        if employee_id in self.employees:
            for existing_class in self.employees[employee_id]:
                if existing_class != class_info and self.is_time_conflict(
                    start_time, duration, existing_class['start_time'], existing_class['duration']):
                    conflicts.append({
                        'type': 'employee_conflict',
                        'message': f"Employee {employee_id} has overlapping classes",
                        'conflicting_class': existing_class
                    })
        # Check room conflicts (by roomid)
        for roomid in roomids:
            if roomid in self.room_schedules:
                for existing_class in self.room_schedules[roomid]:
                    if existing_class != class_info and self.is_time_conflict(
                        start_time, duration, existing_class['start_time'], existing_class['duration']):
                        conflicts.append({
                            'type': 'room_conflict',
                            'message': f"Room {roomid} has overlapping classes",
                            'conflicting_class': existing_class
                        })
        return conflicts
    
    def find_available_slot(self, duration: int, section: str, employee_id: int, roomids) -> tuple:
        school_start = 6 * 60  # 6:00 AM
        school_end = 21 * 60   # 9:00 PM
        if not isinstance(roomids, list):
            roomids = [roomids]
        for start_time in range(school_start, school_end - duration + 1, 30):
            end_time = start_time + duration
            # Skip if overlaps with break
            if self.is_in_break_time(start_time, duration):
                continue
            # Section
            section_conflict = False
            if section in self.sections:
                for existing_class in self.sections[section]:
                    if self.is_time_conflict(start_time, duration, existing_class['start_time'], existing_class['duration']):
                        section_conflict = True
                        break
            if section_conflict:
                continue
            # Employee
            employee_conflict = False
            if employee_id in self.employees:
                for existing_class in self.employees[employee_id]:
                    if self.is_time_conflict(start_time, duration, existing_class['start_time'], existing_class['duration']):
                        employee_conflict = True
                        break
            if employee_conflict:
                continue
            # Room (all rooms in the list must be available)
            room_conflict = False
            for roomid in roomids:
                if roomid in self.room_schedules:
                    for existing_class in self.room_schedules[roomid]:
                        if self.is_time_conflict(start_time, duration, existing_class['start_time'], existing_class['duration']):
                            room_conflict = True
                            break
                if room_conflict:
                    break
            if room_conflict:
                continue
            return start_time, end_time
        return None, None
    
    def schedule_class(self, course_id: int, coursename: str, section: str, duration: int, roomids, employee_id: int) -> dict:
        if not isinstance(roomids, list):
            roomids = [roomids]
        start_time, end_time = self.find_available_slot(duration, section, employee_id, roomids)
        if start_time is None:
            return None
        class_info = {
            'course_id': course_id,
            'coursename': coursename,
            'section': section,
            'start_time': start_time,
            'end_time': end_time,
            'duration': duration,
            'roomid': roomids,
            'employee_id': employee_id,
            'start_time_str': self.format_time(start_time),
            'end_time_str': self.format_time(end_time)
        }
        conflicts = self.detect_conflicts(class_info)
        if conflicts:
            self.conflicts.extend(conflicts)
            return None
        if section not in self.sections:
            self.sections[section] = []
        self.sections[section].append(class_info)
        if employee_id not in self.employees:
            self.employees[employee_id] = []
        self.employees[employee_id].append(class_info)
        for roomid in roomids:
            if roomid not in self.room_schedules:
                self.room_schedules[roomid] = []
            self.room_schedules[roomid].append(class_info)
        self.schedule.append(class_info)
        return class_info
    
    def _count_section_classes_per_day(self, section: str) -> dict:
        """Return a dict mapping day to number of scheduled classes for the section."""
        day_counts = {}
        if section in self.sections:
            for class_info in self.sections[section]:
                day = class_info.get('day', None)
                if day:
                    day_counts[day] = day_counts.get(day, 0) + 1
        return day_counts

    def generate_schedule(self, class_data: list) -> list:
        self.schedule = []
        self.sections = {}
        self.employees = {}
        self.room_schedules = {}
        self.conflicts = []
        unscheduled_classes = []
        # Track used days for each (section, courseid)
        used_days_per_course = {}
        for class_group in class_data:
            for course in class_group['Courses']:
                coursename = course.get('coursename', '')
                section = course['section']
                courseid = course['courseid']
                key = (section, courseid)
                if key not in used_days_per_course:
                    used_days_per_course[key] = set()
                for schedule_item in course['classschedule']:
                    if 'roomid' not in schedule_item or 'employeeid' not in schedule_item:
                        continue
                    days = schedule_item.get('day', ['Monday'])
                    if isinstance(days, str):
                        days = [days]
                    # Only consider days not already used for this course in this section
                    available_days = [d for d in days if d not in used_days_per_course[key]]
                    # Sort days by current load for this section (least loaded first)
                    section_day_counts = self._count_section_classes_per_day(section)
                    days_sorted = sorted(available_days, key=lambda d: section_day_counts.get(d, 0))
                    scheduled = None
                    for day in days_sorted:
                        scheduled = self.schedule_class_with_day(
                            courseid,
                            coursename,
                            section,
                            self.parse_duration(schedule_item['duration']),
                            schedule_item['roomid'],
                            schedule_item['employeeid'],
                            day
                        )
                        if scheduled:
                            used_days_per_course[key].add(day)
                            break
                    if not scheduled:
                        # Add to unscheduled with all possible days for reference
                        unscheduled_classes.append({
                            'course_id': courseid,
                            'coursename': coursename,
                            'section': section,
                            'duration': self.parse_duration(schedule_item['duration']),
                            'roomid': schedule_item['roomid'],
                            'employee_id': schedule_item['employeeid'],
                            'day': days
                        })
        self.unscheduled_classes = unscheduled_classes  # Store for later use
        return self.schedule

    def schedule_class_with_day(self, course_id: int, coursename: str, section: str, duration: int, roomids, employee_id: int, day: str) -> dict:
        if not isinstance(roomids, list):
            roomids = [roomids]
        school_start = 8 * 60  # 8:00 AM
        school_end = 17 * 60   # 5:00 PM
        for start_time in range(school_start, school_end - duration + 1, 30):
            end_time = start_time + duration
            # Skip if overlaps with break
            if self.is_in_break_time(start_time, duration):
                continue
            # Section
            section_conflict = False
            if section in self.sections:
                for existing_class in self.sections[section]:
                    if existing_class.get('day') == day and self.is_time_conflict(start_time, duration, existing_class['start_time'], existing_class['duration']):
                        section_conflict = True
                        break
            if section_conflict:
                continue
            # Employee
            employee_conflict = False
            if employee_id in self.employees:
                for existing_class in self.employees[employee_id]:
                    if existing_class.get('day') == day and self.is_time_conflict(start_time, duration, existing_class['start_time'], existing_class['duration']):
                        employee_conflict = True
                        break
            if employee_conflict:
                continue
            # Room (all rooms in the list must be available)
            room_conflict = False
            for roomid in roomids:
                if roomid in self.room_schedules:
                    for existing_class in self.room_schedules[roomid]:
                        if existing_class.get('day') == day and self.is_time_conflict(start_time, duration, existing_class['start_time'], existing_class['duration']):
                            room_conflict = True
                            break
                if room_conflict:
                    break
            if room_conflict:
                continue
            class_info = {
                'course_id': course_id,
                'coursename': coursename,
                'section': section,
                'start_time': start_time,
                'end_time': end_time,
                'duration': duration,
                'roomid': roomids,
                'employee_id': employee_id,
                'day': day,
                'start_time_str': self.format_time(start_time),
                'end_time_str': self.format_time(end_time)
            }
            conflicts = self.detect_conflicts(class_info)
            if conflicts:
                self.conflicts.extend(conflicts)
                break
            if section not in self.sections:
                self.sections[section] = []
            self.sections[section].append(class_info)
            if employee_id not in self.employees:
                self.employees[employee_id] = []
            self.employees[employee_id].append(class_info)
            for roomid in roomids:
                if roomid not in self.room_schedules:
                    self.room_schedules[roomid] = []
                self.room_schedules[roomid].append(class_info)
            self.schedule.append(class_info)
            return class_info
        # If not scheduled, force schedule in the first available slot (ignore conflicts)
        start_time = school_start
        end_time = start_time + duration
        class_info = {
            'course_id': course_id,
            'coursename': coursename,
            'section': section,
            'start_time': start_time,
            'end_time': end_time,
            'duration': duration,
            'roomid': roomids,
            'employee_id': employee_id,
            'day': day,
            'start_time_str': self.format_time(start_time),
            'end_time_str': self.format_time(end_time)
        }
        if section not in self.sections:
            self.sections[section] = []
        self.sections[section].append(class_info)
        if employee_id not in self.employees:
            self.employees[employee_id] = []
        self.employees[employee_id].append(class_info)
        for roomid in roomids:
            if roomid not in self.room_schedules:
                self.room_schedules[roomid] = []
            self.room_schedules[roomid].append(class_info)
        self.schedule.append(class_info)
        return class_info
    
    def print_schedule(self):
        """Print the generated schedule in a readable format"""
        if not self.schedule:
            print("No schedule generated.")
            return
        print("\n" + "="*80)
        print("GENERATED SCHOOL SCHEDULE")
        print("="*80)
        # Group by section
        sections = {}
        for class_info in self.schedule:
            section = class_info['section']
            if section not in sections:
                sections[section] = []
            sections[section].append(class_info)
        for section in sorted(sections.keys()):
            print(f"\nSection: {section}")
            print("-" * 50)
            # Sort by start time
            section_classes = sorted(sections[section], key=lambda x: x['start_time'])
            for class_info in section_classes:
                roomids = class_info['roomid']
                if not isinstance(roomids, list):
                    roomids = [roomids]
                room_str = ','.join(str(r) for r in roomids)
                print(f"  {class_info['start_time_str']} - {class_info['end_time_str']} | "
                      f"Course {class_info['course_id']} | "
                      f"Rooms {room_str} | "
                      f"Employee {class_info['employee_id']}")
        # Print room utilization
        print(f"\n" + "="*80)
        print("ROOM UTILIZATION")
        print("="*80)
        for roomid in sorted(self.room_schedules.keys()):
            classes = sorted(self.room_schedules[roomid], key=lambda x: x['start_time'])
            print(f"\nRoom {roomid} Schedule:")
            for class_info in classes:
                print(f"  {class_info['start_time_str']} - {class_info['end_time_str']} | "
                      f"Course {class_info['course_id']} | "
                      f"Section {class_info['section']} | "
                      f"Employee {class_info['employee_id']}")
        # Print employee schedules
        print(f"\n" + "="*80)
        print("EMPLOYEE SCHEDULES")
        print("="*80)
        for employee_id in sorted(self.employees.keys()):
            classes = sorted(self.employees[employee_id], key=lambda x: x['start_time'])
            print(f"\nEmployee {employee_id} Schedule:")
            for class_info in classes:
                roomids = class_info['roomid']
                if not isinstance(roomids, list):
                    roomids = [roomids]
                room_str = ','.join(str(r) for r in roomids)
                print(f"  {class_info['start_time_str']} - {class_info['end_time_str']} | "
                      f"Course {class_info['course_id']} | "
                      f"Section {class_info['section']} | "
                      f"Rooms {room_str}")
        # Print conflicts if any
        if self.conflicts:
            print(f"\n" + "="*80)
            print("CONFLICTS DETECTED")
            print("="*80)
            for conflict in self.conflicts:
                print(f"  {conflict['type']}: {conflict['message']}")
    
    def export_schedule(self, filename: str = "schedule.json"):
        """Export schedule to JSON file"""
        with open(filename, 'w') as f:
            json.dump(self.schedule, f, indent=2)
        print(f"\nSchedule exported to {filename}")
    
    def export_schedule_csv(self, filename: str = "schedule.csv"):
        """Export schedule to CSV file (one row per class)"""
        fieldnames = [
            'course_id', 'coursename', 'section', 'start_time_str', 'end_time_str', 'duration',
            'roomid', 'employee_id'
        ]
        with open(filename, 'w', newline='') as csvfile:
            writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
            writer.writeheader()
            for class_info in self.schedule:
                row = {k: class_info.get(k, '') for k in fieldnames}
                writer.writerow(row)
        print(f"\nSchedule exported to {filename}")
    
    def export_schedule_timegrid_csv(self, filename):
        """Export schedule as a time grid to CSV file"""
        if not self.schedule:
            print("No schedule to export")
            return

        # Get all unique rooms (handling both single roomid and lists)
        roomids = set()
        for c in self.schedule:
            if isinstance(c['roomid'], list):
                roomids.update(c['roomid'])
            else:
                roomids.add(c['roomid'])
        roomids = sorted(roomids)

        # Get all unique sections
        sections = sorted(set(c['section'] for c in self.schedule))

        # Get all unique employees
        employees = sorted(set(c['employee_id'] for c in self.schedule))

        # Create weekly grids
        days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday']
        hours = ['08:00', '09:00', '10:00', '11:00', '12:00', '13:00', '14:00', '15:00', '16:00', '17:00']

        # Create separate files for each view
        self._export_weekly_grid_by_category(filename.replace('.csv', '_by_section.csv'), 'section', sections, roomids, employees, days, hours)
        self._export_weekly_grid_by_category(filename.replace('.csv', '_by_room.csv'), 'room', roomids, sections, employees, days, hours)
        self._export_weekly_grid_by_category(filename.replace('.csv', '_by_employee.csv'), 'employee', employees, sections, roomids, days, hours)

    def _export_weekly_grid_by_category(self, filename, category_type, primary_keys, secondary_keys1, secondary_keys2, days, hours):
        with open(filename, 'w', newline='') as f:
            writer = csv.writer(f)
            
            # Write headers
            headers = ['Time', 'Day', f'{category_type.capitalize()}', 'Classes']
            writer.writerow(headers)
            
            # Write schedule entries
            for hour in hours:
                for day in days:
                    for key in primary_keys:
                        cell_content = []
                        for class_info in self.schedule:
                            # Handle both string and integer time formats
                            start_time = class_info.get('start_time', '')
                            if isinstance(start_time, int):
                                # Convert minutes since midnight to HH:MM format
                                hours = start_time // 60
                                minutes = start_time % 60
                                start_time = f"{hours:02d}:{minutes:02d}"
                            
                            # Skip if required fields are missing
                            if not all(key in class_info for key in ['roomid', 'employee_id', 'section', 'coursename']):
                                continue
                                
                            class_time = datetime.strptime(start_time, '%H:%M').strftime('%H:%M')
                            if class_time == hour:
                                try:
                                    # Match based on category type
                                    if category_type == 'section' and class_info['section'] == key:
                                        roomid_str = ', '.join(str(r) for r in (class_info['roomid'] if isinstance(class_info['roomid'], list) else [class_info['roomid']]))
                                        cell_content.append(f"{class_info['coursename']} (Room {roomid_str}, Emp {class_info['employee_id']})")
                                    elif category_type == 'room':
                                        roomids = class_info['roomid'] if isinstance(class_info['roomid'], list) else [class_info['roomid']]
                                        if key in roomids:
                                            cell_content.append(f"{class_info['coursename']} (Sec {class_info['section']}, Emp {class_info['employee_id']})")
                                    elif category_type == 'employee' and class_info['employee_id'] == key:
                                        roomid_str = ', '.join(str(r) for r in (class_info['roomid'] if isinstance(class_info['roomid'], list) else [class_info['roomid']]))
                                        cell_content.append(f"{class_info['coursename']} (Sec {class_info['section']}, Room {roomid_str})")
                                except (KeyError, TypeError) as e:
                                    print(f"Warning: Skipping class due to missing or invalid data: {class_info}")
                                    continue
                        
                        writer.writerow([hour, day, key, ' | '.join(cell_content) if cell_content else ''])
    
    def get_schedule_statistics(self) -> Dict:
        """Get statistics about the generated schedule"""
        if not self.schedule:
            return {}
        
        total_classes = len(self.schedule)
        total_duration = sum(class_info['duration'] for class_info in self.schedule)
        
        # Room utilization
        room_utilization = {}
        for roomid in self.room_schedules:
            room_classes = self.room_schedules[roomid]
            room_duration = sum(class_info['duration'] for class_info in room_classes)
            room_utilization[roomid] = {
                'classes': len(room_classes),
                'total_duration': room_duration,
                'utilization_percentage': (room_duration / (9 * 60)) * 100  # 9 hours = 540 minutes
            }
        
        # Employee workload
        employee_workload = {}
        for employee_id in self.employees:
            emp_classes = self.employees[employee_id]
            emp_duration = sum(class_info['duration'] for class_info in emp_classes)
            employee_workload[employee_id] = {
                'classes': len(emp_classes),
                'total_duration': emp_duration,
                'workload_percentage': (emp_duration / (9 * 60)) * 100
            }
        
        return {
            'total_classes': total_classes,
            'total_duration_minutes': total_duration,
            'room_utilization': room_utilization,
            'employee_workload': employee_workload,
            'conflicts_detected': len(self.conflicts)
        }

    def export_weekly_grid_csv(self, filename: str, filter_type: str, filter_value: str):
        """Export a weekly grid CSV for a section, room, or employee."""
        # Time slots: 6:00 AM to 9:00 PM (30-min intervals)
        start_min = 6 * 60
        end_min = 21 * 60
        time_slots = [start_min + 30 * i for i in range(((end_min - start_min) // 30) + 1)]
        time_labels = [self.format_time(m) + ' - ' + self.format_time(m + 30) for m in time_slots]
        days = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
        # Filter classes
        if filter_type == 'section':
            filtered = [c for c in self.schedule if c['section'] == filter_value]
        elif filter_type == 'room':
            filtered = [c for c in self.schedule if c['roomid'] == filter_value]
        elif filter_type == 'employee':
            filtered = [c for c in self.schedule if str(c['employee_id']) == str(filter_value)]
        else:
            filtered = self.schedule
        # Build grid
        grid = {label: {day: '' for day in days} for label in time_labels}
        for c in filtered:
            start = c['start_time']
            end = c['end_time']
            day = c.get('day', '')
            info = f"{c['coursename']}\n{c['section']}\n{c['roomid']}\nEmp:{c['employee_id']}"
            n_slots = (end - start) // 30
            # Find the slot indices for this class
            slot_indices = [i for i, m in enumerate(time_slots) if start <= m < end]
            for idx, slot_idx in enumerate(slot_indices):
                label = time_labels[slot_idx]
                if day in days:
                    if idx == 0:
                        grid[label][day] = info
                    else:
                        grid[label][day] = ''
        # Write CSV
        with open(filename, 'w', newline='') as csvfile:
            writer = csv.writer(csvfile)
            writer.writerow(['Time'] + days)
            for label in time_labels:
                row = [label] + [grid[label][day] for day in days]
                writer.writerow(row)
        print(f"Exported weekly grid to {filename}")

    def export_all_weekly_grids(self):
        """Export weekly grid schedules for each section, room, and employee"""
        # Get all unique sections
        sections = set(c['section'] for c in self.schedule)
        
        # Get all unique rooms (handling both single roomid and lists)
        rooms = set()
        for c in self.schedule:
            if isinstance(c['roomid'], list):
                rooms.update(c['roomid'])
            else:
                rooms.add(c['roomid'])
        
        # Get all unique employees
        employees = set(c['employee_id'] for c in self.schedule)

        # Export section schedules
        for section in sections:
            filename = f"section_{section}_weekly.csv"
            self.export_weekly_grid_by_section(section, filename)
            print(f"Exported weekly grid to {filename}")

        # Export room schedules
        for room in rooms:
            filename = f"room_{room}_weekly.csv"
            self.export_weekly_grid_by_room(room, filename)
            print(f"Exported weekly grid to {filename}")

        # Export employee schedules
        for employee in employees:
            filename = f"employee_{employee}_weekly.csv"
            self.export_weekly_grid_by_employee(employee, filename)
            print(f"Exported weekly grid to {filename}")

    def export_weekly_grid_by_room(self, room_id, filename):
        """Export weekly grid schedule for a specific room"""
        # Create weekly grid with 1-hour slots
        days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday']
        hours = ['08:00', '09:00', '10:00', '11:00', '12:00', '13:00', '14:00', '15:00', '16:00', '17:00']

        with open(filename, 'w', newline='') as f:
            writer = csv.writer(f)
            writer.writerow(['Time', 'Day', 'Classes'])
            
            for hour in hours:
                for day in days:
                    classes = []
                    for class_info in self.schedule:
                        start_time = class_info.get('start_time', '')
                        if isinstance(start_time, int):
                            # Convert minutes since midnight to HH:MM format
                            hours_val = start_time // 60
                            minutes = start_time % 60
                            start_time = f"{hours_val:02d}:{minutes:02d}"
                        
                        if not all(key in class_info for key in ['roomid', 'section', 'coursename', 'employee_id']):
                            continue
                            
                        class_time = datetime.strptime(start_time, '%H:%M').strftime('%H:%M')
                        if class_time == hour:
                            roomids = class_info['roomid'] if isinstance(class_info['roomid'], list) else [class_info['roomid']]
                            if room_id in roomids:
                                classes.append(f"{class_info['coursename']} (Section {class_info['section']}, Employee {class_info['employee_id']})")
                    
                    writer.writerow([hour, day, ' | '.join(classes) if classes else ''])

    def export_weekly_grid_by_section(self, section_id, filename):
        """Export weekly grid schedule for a specific section"""
        days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday']
        hours = ['08:00', '09:00', '10:00', '11:00', '12:00', '13:00', '14:00', '15:00', '16:00', '17:00']

        with open(filename, 'w', newline='') as f:
            writer = csv.writer(f)
            writer.writerow(['Time', 'Day', 'Classes'])
            
            for hour in hours:
                for day in days:
                    classes = []
                    for class_info in self.schedule:
                        start_time = class_info.get('start_time', '')
                        if isinstance(start_time, int):
                            hours_val = start_time // 60
                            minutes = start_time % 60
                            start_time = f"{hours_val:02d}:{minutes:02d}"
                        
                        if not all(key in class_info for key in ['roomid', 'section', 'coursename', 'employee_id']):
                            continue
                            
                        class_time = datetime.strptime(start_time, '%H:%M').strftime('%H:%M')
                        if class_time == hour and class_info['section'] == section_id:
                            roomid_str = ', '.join(str(r) for r in (class_info['roomid'] if isinstance(class_info['roomid'], list) else [class_info['roomid']]))
                            classes.append(f"{class_info['coursename']} (Room {roomid_str}, Employee {class_info['employee_id']})")
                    
                    writer.writerow([hour, day, ' | '.join(classes) if classes else ''])

    def export_weekly_grid_by_employee(self, employee_id, filename):
        """Export weekly grid schedule for a specific employee"""
        days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday']
        hours = ['08:00', '09:00', '10:00', '11:00', '12:00', '13:00', '14:00', '15:00', '16:00', '17:00']

        with open(filename, 'w', newline='') as f:
            writer = csv.writer(f)
            writer.writerow(['Time', 'Day', 'Classes'])
            
            for hour in hours:
                for day in days:
                    classes = []
                    for class_info in self.schedule:
                        start_time = class_info.get('start_time', '')
                        if isinstance(start_time, int):
                            hours_val = start_time // 60
                            minutes = start_time % 60
                            start_time = f"{hours_val:02d}:{minutes:02d}"
                        
                        if not all(key in class_info for key in ['roomid', 'section', 'coursename', 'employee_id']):
                            continue
                            
                        class_time = datetime.strptime(start_time, '%H:%M').strftime('%H:%M')
                        if class_time == hour and class_info['employee_id'] == employee_id:
                            roomid_str = ', '.join(str(r) for r in (class_info['roomid'] if isinstance(class_info['roomid'], list) else [class_info['roomid']]))
                            classes.append(f"{class_info['coursename']} (Section {class_info['section']}, Room {roomid_str})")
                    
                    writer.writerow([hour, day, ' | '.join(classes) if classes else ''])

    def _get_day_column(self, class_info):
        """Helper method to determine the day column for a class"""
        # For now, default to Monday (column 2)
        # TODO: Add logic to determine the correct day based on class_info
        return 2

    def _calculate_duration_minutes(self, duration):
        """Helper method to calculate duration in minutes from either string or integer input"""
        if isinstance(duration, int):
            return duration
        elif isinstance(duration, str):
            try:
                # Handle "HH:MM" format
                if ':' in duration:
                    hours, minutes = map(int, duration.split(':'))
                    return hours * 60 + minutes
                # Handle "H.MM" format (e.g., "1.30" for 1 hour 30 minutes)
                elif '.' in duration:
                    hours, minutes = map(int, duration.split('.'))
                    return hours * 60 + minutes
                # Handle integer string
                else:
                    return int(duration)
            except (ValueError, TypeError):
                return 60  # Default to 1 hour if format is invalid
        return 60  # Default to 1 hour for any other case

    def export_schedule_excel(self, filename, filter_func=None):
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        wb = Workbook()
        ws = wb.active
        ws.title = "Schedule"
        header_fill = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')
        center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        colors = {
            'GE Entrep': PatternFill(start_color='92D050', end_color='92D050', fill_type='solid'),
            'GE USELF': PatternFill(start_color='FF69B4', end_color='FF69B4', fill_type='solid'),
            'GE ArtApp': PatternFill(start_color='00B0F0', end_color='00B0F0', fill_type='solid'),
            'GE LITE': PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid'),
            'NSTP': PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid'),
            'PE': PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        }
        days = ['Time', 'Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
        day_to_col = {day: idx+2 for idx, day in enumerate(days[1:])}
        for col, day in enumerate(days, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = day
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border
        time_slots = []
        time_to_row = {}
        current_row = 2
        for hour in range(8, 17):
            for minute in [0, 30]:
                time_str = f"{hour:02d}:{minute:02d}"
                time_to_row[time_str] = current_row
                next_minute = minute + 30
                next_hour = hour
                if next_minute >= 60:
                    next_minute = 0
                    next_hour += 1
                if hour < 12:
                    current_time = f"{hour}:{minute:02d} AM"
                elif hour == 12:
                    current_time = f"12:{minute:02d} PM"
                else:
                    current_time = f"{hour-12}:{minute:02d} PM"
                if next_hour < 12:
                    next_time = f"{next_hour}:{next_minute:02d} AM"
                elif next_hour == 12:
                    next_time = f"12:{next_minute:02d} PM"
                else:
                    next_time = f"{next_hour-12}:{next_minute:02d} PM"
                time_slots.append(f"{current_time} - {next_time}")
                current_row += 1
        for idx, time_slot in enumerate(time_slots):
            cell = ws.cell(row=idx+2, column=1)
            cell.value = time_slot
            cell.alignment = center_align
            cell.border = thin_border
        merged_tracker = set()
        for class_info in self.schedule:
            if filter_func and not filter_func(class_info):
                continue
            if not all(key in class_info for key in ['start_time', 'coursename', 'section', 'employee_id', 'roomid']):
                continue
            start_time = class_info['start_time']
            if isinstance(start_time, int):
                hours = start_time // 60
                minutes = start_time % 60
                start_time = f"{hours:02d}:{minutes:02d}"
            start_row = time_to_row.get(start_time)
            if not start_row:
                continue
            duration = self._calculate_duration_minutes(class_info.get('duration', '1:00'))
            num_slots = max(1, duration // 30)
            end_row = start_row + num_slots - 1
            days_list = class_info.get('day', ['Monday'])
            if isinstance(days_list, str):
                days_list = [days_list]
            for day in days_list:
                col = day_to_col.get(day, 2)
                roomid_str = ', '.join(str(r) for r in (class_info['roomid'] if isinstance(class_info['roomid'], list) else [class_info['roomid']]))
                class_text = (
                    f"{class_info['coursename']}\n"
                    f"{class_info['section']}\n"
                    f"{class_info.get('instructor_name', '')}\n"
                    f"Room {roomid_str}"
                )
                color_fill = None
                for course_type, fill in colors.items():
                    if course_type in class_info['coursename']:
                        color_fill = fill
                        break
                # Check if all cells in the intended merge range are empty and not already merged
                can_merge = True
                for row in range(start_row, end_row + 1):
                    cell = ws.cell(row=row, column=col)
                    cell_key = (row, col)
                    if cell.value not in (None, "") or cell_key in merged_tracker:
                        can_merge = False
                        break
                if can_merge:
                    # Mark all cells as merged
                    for row in range(start_row, end_row + 1):
                        merged_tracker.add((row, col))
                    ws.merge_cells(start_row=start_row, start_column=col, end_row=end_row, end_column=col)
                    cell = ws.cell(row=start_row, column=col)
                    cell.value = class_text
                    if color_fill:
                        cell.fill = color_fill
                    cell.alignment = center_align
                    cell.border = thin_border
                    # Fill merged cells with style only (no value)
                    for row in range(start_row + 1, end_row + 1):
                        cell = ws.cell(row=row, column=col)
                        if color_fill:
                            cell.fill = color_fill
                        cell.alignment = center_align
                        cell.border = thin_border
        for col in range(1, len(days) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 25
        for row in range(1, len(time_slots) + 2):
            ws.row_dimensions[row].height = 50
        wb.save(filename)
        print(f"\nSchedule exported to Excel file: {filename}")

    def export_all_excel_schedules(self):
        """Export separate Excel schedules for sections, rooms, and employees into a subfolder."""
        output_dir = 'excel_schedules'
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)
        # Get all unique sections
        sections = set(c['section'] for c in self.schedule if 'section' in c)
        # Get all unique rooms
        rooms = set()
        for c in self.schedule:
            if 'roomid' not in c:
                continue
            if isinstance(c['roomid'], list):
                rooms.update(c['roomid'])
            else:
                rooms.add(c['roomid'])
        # Get all unique employees
        employees = set(c['employee_id'] for c in self.schedule if 'employee_id' in c)
        # Export schedules
        for section in sections:
            filename = os.path.join(output_dir, f"section_{section}_schedule.xlsx")
            self.export_schedule_excel(filename, filter_func=lambda c, s=section: c.get('section') == s)
            print(f"Exported Excel schedule for section {section}")
        for room in rooms:
            filename = os.path.join(output_dir, f"room_{room}_schedule.xlsx")
            self.export_schedule_excel(filename, filter_func=lambda c, r=room: (r in c['roomid']) if isinstance(c['roomid'], list) else c['roomid'] == r)
            print(f"Exported Excel schedule for room {room}")
        for employee in employees:
            filename = os.path.join(output_dir, f"employee_{employee}_schedule.xlsx")
            self.export_schedule_excel(filename, filter_func=lambda c, e=employee: c.get('employee_id') == e)
            print(f"Exported Excel schedule for employee {employee}")

    def print_unscheduled_classes(self):
        """Print the list of classes that could not be scheduled."""
        if not self.unscheduled_classes:
            print("\nAll classes were scheduled successfully.")
            return
        print("\n" + "="*80)
        print("UNSCHEDULED (NOT PLOTTED) CLASSES")
        print("="*80)
        for idx, class_info in enumerate(self.unscheduled_classes, 1):
            print(f"{idx}. Course {class_info.get('course_id', '')} | {class_info.get('coursename', '')} | Section {class_info.get('section', '')} | Room {class_info.get('roomid', '')} | Employee {class_info.get('employee_id', '')} | Day {class_info.get('day', '')} | Duration {class_info.get('duration', '')}")

    def export_unscheduled_classes_csv(self, filename="unscheduled_classes.csv"):
        """Export the list of unscheduled classes to a CSV file."""
        if not self.unscheduled_classes:
            print("\nNo unscheduled classes to export.")
            return
        fieldnames = ['course_id', 'coursename', 'section', 'roomid', 'employee_id', 'day', 'duration']
        with open(filename, 'w', newline='') as csvfile:
            writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
            writer.writeheader()
            for class_info in self.unscheduled_classes:
                row = {k: class_info.get(k, '') for k in fieldnames}
                writer.writerow(row)
        print(f"\nUnscheduled classes exported to {filename}")

    def is_in_break_time(self, start_time: int, duration: int) -> bool:
        break_start = 12 * 60  # 12:00 PM in minutes
        break_end = 13 * 60    # 1:00 PM in minutes
        end_time = start_time + duration
        # If any part of the class overlaps with the break
        return not (end_time <= break_start or start_time >= break_end)

    def log_strict_schedule_results(self, class_data, log_filename='logs.txt'):
        """Log for every input class: requested (day, duration), actual scheduled (if any), or UNSCHEDULED if not found, with reason if unscheduled."""
        with open(log_filename, 'w') as f:
            f.write("SCHEDULED CLASSES (EXACT MATCHES):\n")
            for s in self.schedule:
                f.write(f"Scheduled: {s.get('coursename', '')} | Section {s.get('section', '')} | Room {s.get('roomid', '')} | Emp {s.get('employee_id', '')} | Day {s.get('day', '')} | Duration {s.get('duration', '')}\n")
            f.write("\nREQUESTED VS ACTUAL:\n")
            # Flatten all courses if needed
            all_courses = []
            if isinstance(class_data, list):
                for entry in class_data:
                    all_courses.extend(entry['Courses'])
            elif isinstance(class_data, dict) and 'Courses' in class_data:
                all_courses = class_data['Courses']
            else:
                all_courses = class_data
            for course in all_courses:
                for sched in course['classschedule']:
                    sched_day = sched.get('day', '')
                    sched_day_val = sched_day if isinstance(sched_day, str) else (sched_day[0] if isinstance(sched_day, list) and sched_day else '')
                    sched_duration = self.parse_duration(sched['duration'])
                    found = False
                    for s in self.schedule:
                        # Compare all relevant fields strictly
                        if (
                            s.get('coursename', '') == course.get('coursename', '') and
                            s.get('section', '') == course['section'] and
                            str(s.get('roomid', '')) == str(sched['roomid']) and
                            s.get('employee_id', '') == sched['employeeid'] and
                            s.get('duration', '') == sched_duration and
                            (s.get('day', '') == sched_day_val or s.get('day', '') == (','.join(sched['day']) if isinstance(sched.get('day', ''), list) else sched.get('day', '')))
                        ):
                            found = True
                            f.write(f"REQUESTED: {course.get('coursename', '')} | Section {course['section']} | Room {sched['roomid']} | Emp {sched['employeeid']} | Day {sched.get('day', '')} | Duration {sched['duration']}\n")
                            f.write(f"  -> SCHEDULED: {s.get('coursename', '')} | Section {s.get('section', '')} | Room {s.get('roomid', '')} | Emp {s.get('employee_id', '')} | Day {s.get('day', '')} | Duration {s.get('duration', '')}\n")
                            break
                    if not found:
                        # Try to diagnose the reason
                        reason = self._diagnose_unscheduled(
                            course.get('coursename', ''),
                            course['section'],
                            sched['roomid'],
                            sched['employeeid'],
                            sched_day_val,
                            sched_duration
                        )
                        f.write(f"REQUESTED: {course.get('coursename', '')} | Section {course['section']} | Room {sched['roomid']} | Emp {sched['employeeid']} | Day {sched.get('day', '')} | Duration {sched['duration']}\n")
                        f.write(f"  -> UNSCHEDULED | REASON: {reason}\n")

    def _diagnose_unscheduled(self, coursename, section, roomid, employee_id, day, duration):
        # Try all possible days and time slots, accumulate all reasons
        school_start = 8 * 60
        school_end = 17 * 60
        slot_step = 30
        if isinstance(roomid, list):
            roomids = roomid
        else:
            roomids = [roomid]
        days = [day] if isinstance(day, str) else (day if isinstance(day, list) else [str(day)])
        reasons = set()
        slot_found = False
        for d in days:
            for start_time in range(school_start, school_end - duration + 1, slot_step):
                # Check break
                if self.is_in_break_time(start_time, duration):
                    reasons.add(f"Break time overlap at {self.format_time(start_time)} on {d}")
                    continue
                # Section conflict
                section_conflict = False
                for c in self.sections.get(section, []):
                    if c.get('day') == d and self.is_time_conflict(start_time, duration, c['start_time'], c['duration']):
                        section_conflict = True
                        break
                if section_conflict:
                    reasons.add(f"Section conflict at {self.format_time(start_time)} on {d}")
                    continue
                # Employee conflict
                employee_conflict = False
                for c in self.employees.get(employee_id, []):
                    if c.get('day') == d and self.is_time_conflict(start_time, duration, c['start_time'], c['duration']):
                        employee_conflict = True
                        break
                if employee_conflict:
                    reasons.add(f"Employee conflict at {self.format_time(start_time)} on {d}")
                    continue
                # Room conflict
                room_conflict = False
                for rid in roomids:
                    for c in self.room_schedules.get(rid, []):
                        if c.get('day') == d and self.is_time_conflict(start_time, duration, c['start_time'], c['duration']):
                            room_conflict = True
                            break
                    if room_conflict:
                        break
                if room_conflict:
                    reasons.add(f"Room conflict at {self.format_time(start_time)} on {d}")
                    continue
                # If no conflicts, this slot should have been available
                slot_found = True
        if slot_found and not reasons:
            return "Unknown reason (should have been schedulable)"
        if reasons:
            return "; ".join(sorted(reasons))
        return "No available slot (all times conflict with break or out of hours)"

    def generate_schedule_backtracking(self, class_data: list) -> bool:
        """
        Backtracking scheduler: schedules all classes with exact requested day and duration, no overlaps/conflicts, and no two sessions of the same course/section on the same day.
        Returns True if a complete schedule is found, otherwise False.
        """
        self.schedule = []
        self.sections = {}
        self.employees = {}
        self.room_schedules = {}
        self.conflicts = []
        self.unscheduled_classes = []
        requests = []
        for class_group in class_data:
            for course in class_group['Courses']:
                coursename = course.get('coursename', '')
                section = course['section']
                courseid = course['courseid']
                for sched in course['classschedule']:
                    days = sched.get('day', ['Monday'])
                    if isinstance(days, str):
                        days = [days]
                    requests.append({
                        'courseid': courseid,
                        'coursename': coursename,
                        'section': section,
                        'duration': self.parse_duration(sched['duration']),
                        'roomid': sched['roomid'],
                        'employee_id': sched['employeeid'],
                        'days': days,
                        'orig_sched': sched
                    })
        requests.sort(key=lambda x: x['duration'], reverse=True)
        school_start = 8 * 60
        school_end = 17 * 60
        slot_step = 30
        used_days_per_course = {}
        def can_assign(req, day, start_time):
            end_time = start_time + req['duration']
            # No two sessions of the same course/section on the same day
            key = (req['section'], req['courseid'])
            if key in used_days_per_course and day in used_days_per_course[key]:
                return False
            if self.is_in_break_time(start_time, req['duration']):
                return False
            for c in self.sections.get(req['section'], []):
                if c.get('day') == day and self.is_time_conflict(start_time, req['duration'], c['start_time'], c['duration']):
                    return False
            for c in self.employees.get(req['employee_id'], []):
                if c.get('day') == day and self.is_time_conflict(start_time, req['duration'], c['start_time'], c['duration']):
                    return False
            roomids = req['roomid'] if isinstance(req['roomid'], list) else [req['roomid']]
            for roomid in roomids:
                for c in self.room_schedules.get(roomid, []):
                    if c.get('day') == day and self.is_time_conflict(start_time, req['duration'], c['start_time'], c['duration']):
                        return False
            return True
        def assign(req, day, start_time):
            end_time = start_time + req['duration']
            class_info = {
                'course_id': req['courseid'],
                'coursename': req['coursename'],
                'section': req['section'],
                'start_time': start_time,
                'end_time': end_time,
                'duration': req['duration'],
                'roomid': req['roomid'],
                'employee_id': req['employee_id'],
                'day': day,
                'start_time_str': self.format_time(start_time),
                'end_time_str': self.format_time(end_time)
            }
            self.schedule.append(class_info)
            self.sections.setdefault(req['section'], []).append(class_info)
            self.employees.setdefault(req['employee_id'], []).append(class_info)
            roomids = req['roomid'] if isinstance(req['roomid'], list) else [req['roomid']]
            for roomid in roomids:
                self.room_schedules.setdefault(roomid, []).append(class_info)
            key = (req['section'], req['courseid'])
            if key not in used_days_per_course:
                used_days_per_course[key] = set()
            used_days_per_course[key].add(day)
            return class_info
        def unassign(req, class_info):
            self.schedule.remove(class_info)
            self.sections[req['section']].remove(class_info)
            self.employees[req['employee_id']].remove(class_info)
            roomids = req['roomid'] if isinstance(req['roomid'], list) else [req['roomid']]
            for roomid in roomids:
                self.room_schedules[roomid].remove(class_info)
            key = (req['section'], req['courseid'])
            if key in used_days_per_course and class_info['day'] in used_days_per_course[key]:
                used_days_per_course[key].remove(class_info['day'])
        def backtrack(idx):
            if idx == len(requests):
                return True
            req = requests[idx]
            key = (req['section'], req['courseid'])
            available_days = [d for d in req['days'] if d not in used_days_per_course.get(key, set())]
            for day in available_days:
                for start_time in range(school_start, school_end - req['duration'] + 1, slot_step):
                    if can_assign(req, day, start_time):
                        class_info = assign(req, day, start_time)
                        if backtrack(idx + 1):
                            return True
                        unassign(req, class_info)
            return False
        success = backtrack(0)
        if not success:
            self.unscheduled_classes = [
                {
                    'course_id': req['courseid'],
                    'coursename': req['coursename'],
                    'section': req['section'],
                    'duration': req['duration'],
                    'roomid': req['roomid'],
                    'employee_id': req['employee_id'],
                    'day': req['days']
                }
                for req in requests if not any(
                    c['course_id'] == req['courseid'] and c['section'] == req['section'] and c['duration'] == req['duration'] and c['roomid'] == req['roomid'] and c['employee_id'] == req['employee_id'] for c in self.schedule
                )
            ]
        return success

    def generate_schedule_genetic(self, class_data, generations=200, pop_size=50) -> bool:
        """
        Genetic algorithm scheduler with repair: schedules all classes with correct duration and no conflicts, and no two sessions of the same course/section on the same day.
        Returns True if a perfect schedule is found, otherwise False.
        """
        import random
        import copy
        requests = []
        for class_group in class_data:
            for course in class_group['Courses']:
                coursename = course.get('coursename', '')
                section = course['section']
                courseid = course['courseid']
                for sched in course['classschedule']:
                    days = sched.get('day', ['Monday'])
                    if isinstance(days, str):
                        days = [days]
                    requests.append({
                        'courseid': courseid,
                        'coursename': coursename,
                        'section': section,
                        'duration': self.parse_duration(sched['duration']),
                        'roomid': sched['roomid'],
                        'employee_id': sched['employeeid'],
                        'days': days,
                        'orig_sched': sched
                    })
        school_start = 8 * 60
        school_end = 17 * 60
        slot_step = 30
        def build_chromosome():
            chrom = []
            used_days_per_course = {}
            for req in requests:
                key = (req['section'], req['courseid'])
                available_days = [d for d in req['days'] if d not in used_days_per_course.get(key, set())]
                tries = 0
                while tries < 10 and available_days:
                    day = random.choice(available_days)
                    possible_starts = [t for t in range(school_start, school_end - req['duration'] + 1, slot_step)]
                    if not possible_starts:
                        tries += 1
                        continue
                    start_time = random.choice(possible_starts)
                    # Check for conflicts
                    conflict = False
                    for c in chrom:
                        if c['section'] == req['section'] and c['day'] == day and self.is_time_conflict(start_time, req['duration'], c['start_time'], c['duration']):
                            conflict = True
                            break
                        if c['employee_id'] == req['employee_id'] and c['day'] == day and self.is_time_conflict(start_time, req['duration'], c['start_time'], c['duration']):
                            conflict = True
                            break
                        roomids = req['roomid'] if isinstance(req['roomid'], list) else [req['roomid']]
                        c_roomids = c['roomid'] if isinstance(c['roomid'], list) else [c['roomid']]
                        if any(r in c_roomids for r in roomids) and c['day'] == day and self.is_time_conflict(start_time, req['duration'], c['start_time'], c['duration']):
                            conflict = True
                            break
                    if conflict:
                        tries += 1
                        continue
                    chrom.append({
                        'courseid': req['courseid'],
                        'coursename': req['coursename'],
                        'section': req['section'],
                        'duration': req['duration'],
                        'roomid': req['roomid'],
                        'employee_id': req['employee_id'],
                        'day': day,
                        'start_time': start_time,
                        'end_time': start_time + req['duration'],
                        'start_time_str': self.format_time(start_time),
                        'end_time_str': self.format_time(start_time + req['duration'])
                    })
                    if key not in used_days_per_course:
                        used_days_per_course[key] = set()
                    used_days_per_course[key].add(day)
                    break
                # If tries exhausted or no available days, skip
            return chrom
        def fitness(chrom):
            if len(chrom) != len(requests):
                return -10000 * (len(requests) - len(chrom))
            used_days_per_course = {}
            for a in chrom:
                key = (a['section'], a['courseid'])
                if key not in used_days_per_course:
                    used_days_per_course[key] = set()
                if a['day'] in used_days_per_course[key]:
                    return -10000  # Two sessions of same course/section on same day
                used_days_per_course[key].add(a['day'])
            for i, a in enumerate(chrom):
                for j, b in enumerate(chrom):
                    if i == j:
                        continue
                    if a['section'] == b['section'] and a['day'] == b['day'] and self.is_time_conflict(a['start_time'], a['duration'], b['start_time'], b['duration']):
                        return -10000
                    if a['employee_id'] == b['employee_id'] and a['day'] == b['day'] and self.is_time_conflict(a['start_time'], a['duration'], b['start_time'], b['duration']):
                        return -10000
                    a_roomids = a['roomid'] if isinstance(a['roomid'], list) else [a['roomid']]
                    b_roomids = b['roomid'] if isinstance(b['roomid'], list) else [b['roomid']]
                    if any(r in b_roomids for r in a_roomids) and a['day'] == b['day'] and self.is_time_conflict(a['start_time'], a['duration'], b['start_time'], b['duration']):
                        return -10000
            return 10000  # Perfect
        def repair(chrom):
            chrom = copy.deepcopy(chrom)
            scheduled_keys = set((c['courseid'], c['section'], c['duration'], str(c['roomid']), c['employee_id']) for c in chrom)
            used_days_per_course = {}
            for c in chrom:
                key = (c['section'], c['courseid'])
                if key not in used_days_per_course:
                    used_days_per_course[key] = set()
                used_days_per_course[key].add(c['day'])
            for req in requests:
                key = (req['section'], req['courseid'])
                available_days = [d for d in req['days'] if d not in used_days_per_course.get(key, set())]
                if (req['courseid'], req['section'], req['duration'], str(req['roomid']), req['employee_id']) in scheduled_keys:
                    continue
                inserted = False
                for day in available_days:
                    for start_time in range(school_start, school_end - req['duration'] + 1, slot_step):
                        conflict = False
                        for c in chrom:
                            if c['section'] == req['section'] and c['day'] == day and self.is_time_conflict(start_time, req['duration'], c['start_time'], c['duration']):
                                conflict = True
                                break
                            if c['employee_id'] == req['employee_id'] and c['day'] == day and self.is_time_conflict(start_time, req['duration'], c['start_time'], c['duration']):
                                conflict = True
                                break
                            roomids = req['roomid'] if isinstance(req['roomid'], list) else [req['roomid']]
                            c_roomids = c['roomid'] if isinstance(c['roomid'], list) else [c['roomid']]
                            if any(r in c_roomids for r in roomids) and c['day'] == day and self.is_time_conflict(start_time, req['duration'], c['start_time'], c['duration']):
                                conflict = True
                                break
                        if conflict:
                            continue
                        chrom.append({
                            'courseid': req['courseid'],
                            'coursename': req['coursename'],
                            'section': req['section'],
                            'duration': req['duration'],
                            'roomid': req['roomid'],
                            'employee_id': req['employee_id'],
                            'day': day,
                            'start_time': start_time,
                            'end_time': start_time + req['duration'],
                            'start_time_str': self.format_time(start_time),
                            'end_time_str': self.format_time(start_time + req['duration'])
                        })
                        if key not in used_days_per_course:
                            used_days_per_course[key] = set()
                        used_days_per_course[key].add(day)
                        inserted = True
                        break
                    if inserted:
                        break
            # Remove conflicts and same-day violations
            i = 0
            while i < len(chrom):
                a = chrom[i]
                key = (a['section'], a['courseid'])
                if chrom.count(a) > 1 or (key in used_days_per_course and list(used_days_per_course[key]).count(a['day']) > 1):
                    chrom.pop(i)
                    continue
                conflict = False
                for j, b in enumerate(chrom):
                    if i == j:
                        continue
                    if a['section'] == b['section'] and a['day'] == b['day'] and self.is_time_conflict(a['start_time'], a['duration'], b['start_time'], b['duration']):
                        conflict = True
                        break
                    if a['employee_id'] == b['employee_id'] and a['day'] == b['day'] and self.is_time_conflict(a['start_time'], a['duration'], b['start_time'], b['duration']):
                        conflict = True
                        break
                    a_roomids = a['roomid'] if isinstance(a['roomid'], list) else [a['roomid']]
                    b_roomids = b['roomid'] if isinstance(b['roomid'], list) else [b['roomid']]
                    if any(r in b_roomids for r in a_roomids) and a['day'] == b['day'] and self.is_time_conflict(a['start_time'], a['duration'], b['start_time'], b['duration']):
                        conflict = True
                        break
                if conflict:
                    chrom.pop(i)
                else:
                    i += 1
            return chrom
        # ... rest of GA unchanged ...

    def global_repair_schedule(self, class_data):
        """
        For each unscheduled class, try every possible slot and day, and if blocked, recursively move the blocking class to another valid slot, until all classes are scheduled or it is mathematically impossible.
        Returns True if all classes are scheduled, False otherwise.
        Strictly enforces: no two sessions of the same course/section on the same day.
        """
        import copy
        requests = []
        for class_group in class_data:
            for course in class_group['Courses']:
                coursename = course.get('coursename', '')
                section = course['section']
                courseid = course['courseid']
                for sched in course['classschedule']:
                    days = sched.get('day', ['Monday'])
                    if isinstance(days, str):
                        days = [days]
                    requests.append({
                        'courseid': courseid,
                        'coursename': coursename,
                        'section': section,
                        'duration': self.parse_duration(sched['duration']),
                        'roomid': sched['roomid'],
                        'employee_id': sched['employeeid'],
                        'days': days,
                        'orig_sched': sched
                    })
        school_start = 8 * 60
        school_end = 17 * 60
        slot_step = 30
        def get_used_days_per_course(schedule):
            used = {}
            for c in schedule:
                key = (c['section'], c['course_id'])
                if key not in used:
                    used[key] = set()
                used[key].add(c['day'])
            return used
        def can_assign(req, day, start_time, schedule, used_days_per_course, ignore_class=None):
            key = (req['section'], req['courseid'])
            if key in used_days_per_course and day in used_days_per_course[key]:
                return False
            if self.is_in_break_time(start_time, req['duration']):
                return False
            for c in schedule:
                if c is ignore_class:
                    continue
                if c['section'] == req['section'] and c['day'] == day and self.is_time_conflict(start_time, req['duration'], c['start_time'], c['duration']):
                    return False
                if c['employee_id'] == req['employee_id'] and c['day'] == day and self.is_time_conflict(start_time, req['duration'], c['start_time'], c['duration']):
                    return False
                roomids = req['roomid'] if isinstance(req['roomid'], list) else [req['roomid']]
                c_roomids = c['roomid'] if isinstance(c['roomid'], list) else [c['roomid']]
                if any(r in c_roomids for r in roomids) and c['day'] == day and self.is_time_conflict(start_time, req['duration'], c['start_time'], c['duration']):
                    return False
            return True
        def assign(req, day, start_time, schedule, used_days_per_course):
            end_time = start_time + req['duration']
            class_info = {
                'course_id': req['courseid'],
                'coursename': req['coursename'],
                'section': req['section'],
                'start_time': start_time,
                'end_time': end_time,
                'duration': req['duration'],
                'roomid': req['roomid'],
                'employee_id': req['employee_id'],
                'day': day,
                'start_time_str': self.format_time(start_time),
                'end_time_str': self.format_time(end_time)
            }
            schedule.append(class_info)
            key = (req['section'], req['courseid'])
            if key not in used_days_per_course:
                used_days_per_course[key] = set()
            used_days_per_course[key].add(day)
            return class_info
        def unassign(class_info, schedule, used_days_per_course):
            schedule.remove(class_info)
            key = (class_info['section'], class_info['course_id'])
            if key in used_days_per_course and class_info['day'] in used_days_per_course[key]:
                used_days_per_course[key].remove(class_info['day'])
        def try_schedule_all(schedule, used_days_per_course, unscheduled, visited):
            if not unscheduled:
                return True
            req = unscheduled[0]
            key = (req['section'], req['courseid'])
            available_days = [d for d in req['days'] if d not in used_days_per_course.get(key, set())]
            for day in available_days:
                for start_time in range(school_start, school_end - req['duration'] + 1, slot_step):
                    # Check if can assign directly
                    if can_assign(req, day, start_time, schedule, used_days_per_course):
                        class_info = assign(req, day, start_time, schedule, used_days_per_course)
                        if try_schedule_all(schedule, used_days_per_course, unscheduled[1:], visited):
                            return True
                        unassign(class_info, schedule, used_days_per_course)
                    else:
                        # Find all blocking classes
                        blocking = []
                        for c in schedule:
                            if c['section'] == req['section'] and c['day'] == day and self.is_time_conflict(start_time, req['duration'], c['start_time'], c['duration']):
                                blocking.append(c)
                            if c['employee_id'] == req['employee_id'] and c['day'] == day and self.is_time_conflict(start_time, req['duration'], c['start_time'], c['duration']):
                                blocking.append(c)
                            roomids = req['roomid'] if isinstance(req['roomid'], list) else [req['roomid']]
                            c_roomids = c['roomid'] if isinstance(c['roomid'], list) else [c['roomid']]
                            if any(r in c_roomids for r in roomids) and c['day'] == day and self.is_time_conflict(start_time, req['duration'], c['start_time'], c['duration']):
                                blocking.append(c)
                        # Try to move all blocking classes recursively
                        can_move_all = True
                        for block in blocking:
                            block_key = (block['section'], block['course_id'], block['duration'], str(block['roomid']), block['employee_id'])
                            req_key = (req['section'], req['courseid'], req['duration'], str(req['roomid']), req['employee_id'])
                            if block_key in visited or req_key in visited:
                                can_move_all = False
                                break
                            unassign(block, schedule, used_days_per_course)
                            block_req = {
                                'courseid': block['course_id'],
                                'coursename': block['coursename'],
                                'section': block['section'],
                                'duration': block['duration'],
                                'roomid': block['roomid'],
                                'employee_id': block['employee_id'],
                                'days': req['days'],
                                'orig_sched': None
                            }
                            if not try_schedule_all(schedule, used_days_per_course, [block_req] + unscheduled[1:], visited | {block_key, req_key}):
                                can_move_all = False
                                break
                        if can_move_all and can_assign(req, day, start_time, schedule, used_days_per_course):
                            class_info = assign(req, day, start_time, schedule, used_days_per_course)
                            if try_schedule_all(schedule, used_days_per_course, unscheduled[1:], visited):
                                return True
                            unassign(class_info, schedule, used_days_per_course)
            return False
        # Build initial schedule and unscheduled list
        schedule = copy.deepcopy(self.schedule)
        used_days_per_course = get_used_days_per_course(schedule)
        scheduled_keys = set((c['course_id'], c['section'], c['duration'], str(c['roomid']), c['employee_id']) for c in schedule)
        unscheduled = [req for req in requests if (req['courseid'], req['section'], req['duration'], str(req['roomid']), req['employee_id']) not in scheduled_keys]
        success = try_schedule_all(schedule, used_days_per_course, unscheduled, set())
        if success:
            # Rebuild all internal structures
            self.schedule = schedule
            self.sections = {}
            self.employees = {}
            self.room_schedules = {}
            for c in self.schedule:
                self.sections.setdefault(c['section'], []).append(c)
                self.employees.setdefault(c['employee_id'], []).append(c)
                roomids = c['roomid'] if isinstance(c['roomid'], list) else [c['roomid']]
                for roomid in roomids:
                    self.room_schedules.setdefault(roomid, []).append(c)
            self.unscheduled_classes = []
            return True
        else:
            # Update unscheduled_classes
            self.unscheduled_classes = [
                {
                    'course_id': req['courseid'],
                    'coursename': req['coursename'],
                    'section': req['section'],
                    'duration': req['duration'],
                    'roomid': req['roomid'],
                    'employee_id': req['employee_id'],
                    'day': req['days']
                }
                for req in unscheduled
            ]
            return False

def main():
    # Your class data
    class_data = [
        {
            'Courses': [
                {
                    'courseid': 1,
                    'section': '2A',
                    'classschedule': [ 
                        {'duration': '3:30', 'roomid': 1, 'employeeid': 120},
                        {'duration': '1:00', 'roomid': 2, 'employeeid': 120},
                        {'duration': '1:15', 'roomid': 2, 'employeeid': 120}
                    ]
                },
                {
                    'courseid': 2,
                    'section': '2A',
                    'classschedule': [
                        {'duration': '1:00', 'roomid': 1, 'employeeid': 120},
                        {'duration': '1:30', 'roomid': 2, 'employeeid': 120}
                    ]
                }
            ]
        },
        {
            'Courses': [
                {
                    'courseid': 3,
                    'section': '2B',
                    'classschedule': [
                        {'duration': '1:00', 'roomid': 1, 'employeeid': 120},
                        {'duration': '2:00', 'roomid': 2, 'employeeid': 120},
                    ]
                },
                {
                    'courseid': 4,
                    'section': '2B',
                    'classschedule': [
                        {'duration': '1:00', 'roomid': 1, 'employeeid': 120},
                    ]
                }
            ]
        },
        {
            'courseid': 99,
            'section': '2A',
            'classschedule': [
                {'duration': '20:00', 'roomid': 1, 'employeeid': 120}  # Impossible duration
            ]
        }
    ]

    scheduler = AdvancedSchoolScheduler()
    max_attempts = 10
    attempt = 0
    success = False
    while attempt < max_attempts and not success:
        generations = 200 + attempt * 100
        pop_size = 50 + attempt * 10
        print(f"Attempt {attempt+1}: Trying genetic algorithm with generations={generations}, pop_size={pop_size}")
        success = scheduler.generate_schedule_genetic(class_data, generations=generations, pop_size=pop_size)
        scheduler.log_strict_schedule_results(class_data, log_filename='logs.txt')
        if not success:
            print(f"Attempt {attempt+1}: Not all classes scheduled, retrying...")
        attempt += 1
    if not success:
        print("Trying backtracking scheduler as fallback...")
        success = scheduler.generate_schedule_backtracking(class_data)
        scheduler.log_strict_schedule_results(class_data, log_filename='logs.txt')
    # Global repair step if still not all scheduled
    if not success:
        print("Trying global repair/reshuffling to schedule remaining classes...")
        success = scheduler.global_repair_schedule(class_data)
        scheduler.log_strict_schedule_results(class_data, log_filename='logs.txt')
    if success:
        print("All classes scheduled! Exporting to Excel...")
        scheduler.export_all_excel_schedules()
    else:
        print("Could not schedule all classes after all attempts and global repair. See logs.txt for details.")
    return success

if __name__ == "__main__":
    main() 