import random
import copy
from typing import List, Dict
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os
import json
from openpyxl.cell.cell import MergedCell

# Helper: parse duration string to minutes
def parse_duration(duration_str):
    if ':' in duration_str:
        hours, minutes = map(int, duration_str.split(':'))
        return hours * 60 + minutes
    else:
        return int(duration_str) * 60

days_of_week = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]
time_slots = [8*60 + 30*i for i in range(18)]  # 8:00 to 17:30, 30-min slots

def random_assignment(class_entry):
    # Pick a random day and random start time (fit duration)
    duration = parse_duration(class_entry['duration'])
    possible_days = class_entry.get('day', days_of_week)
    if isinstance(possible_days, str):
        possible_days = [possible_days]
    day = random.choice(possible_days)
    latest_start = max([t for t in time_slots if t + duration <= 17*60])
    start_time = random.choice([t for t in time_slots if t + duration <= 17*60])
    return {'day': day, 'start_time': start_time}

def build_chromosome(class_data):
    assignments = []
    for course in class_data['Courses']:
        for sched in course['classschedule']:
            assignment = random_assignment(sched)
            assignments.append({
                'courseid': course['courseid'],
                'coursename': course.get('coursename', ''),
                'section': course['section'],
                'roomid': sched['roomid'],
                'employeeid': sched['employeeid'],
                'duration': parse_duration(sched['duration']),
                'assignment': assignment
            })
    return assignments

def fitness(chromosome, class_data=None):
    score = 0
    scheduled = 0
    conflicts = 0
    used_days = set()
    course_day = set()
    for i, a in enumerate(chromosome):
        scheduled += 1
        used_days.add((a['section'], a['assignment']['day']))
        key = (a['section'], a['courseid'], a['assignment']['day'])
        if key in course_day:
            score -= 10
        else:
            course_day.add(key)
        for j, b in enumerate(chromosome):
            if i == j:
                continue
            if a['section'] == b['section'] and a['assignment']['day'] == b['assignment']['day']:
                if is_time_conflict(a, b):
                    score -= 100
                    conflicts += 1
            if a['roomid'] == b['roomid'] and a['assignment']['day'] == b['assignment']['day']:
                if is_time_conflict(a, b):
                    score -= 100
                    conflicts += 1
            if a['employeeid'] == b['employeeid'] and a['assignment']['day'] == b['assignment']['day']:
                if is_time_conflict(a, b):
                    score -= 100
                    conflicts += 1
    score += scheduled
    score += 0.1 * len(used_days)
    if class_data is not None:
        input_keys = set(
            (
                course.get('coursename', ''),
                course['section'],
                str(sched['roomid']),
                sched['employeeid'],
                parse_duration(sched['duration']),
                ','.join(sched['day']) if isinstance(sched.get('day', ''), list) else sched.get('day', '')
            )
            for course in class_data['Courses'] for sched in course['classschedule']
        )
        scheduled_keys = set(
            (
                s['coursename'],
                s['section'],
                str(s['roomid']),
                s['employeeid'],
                s['duration'],
                s['day'] if isinstance(s['day'], str) else ','.join(s['day'])
            )
            for s in chromosome_to_schedule(chromosome)
        )
        missing = input_keys - scheduled_keys
        score -= 200 * len(missing)
    return score

def is_time_conflict(a, b):
    s1, d1 = a['assignment']['start_time'], a['duration']
    s2, d2 = b['assignment']['start_time'], b['duration']
    e1, e2 = s1 + d1, s2 + d2
    return not (e1 <= s2 or e2 <= s1)

def mutate(chromosome):
    # Randomly change the day or start_time of a random assignment
    c = copy.deepcopy(chromosome)
    idx = random.randint(0, len(c)-1)
    if random.random() < 0.5:
        # Change day
        c[idx]['assignment']['day'] = random.choice(days_of_week)
    else:
        # Change time
        duration = c[idx]['duration']
        c[idx]['assignment']['start_time'] = random.choice([t for t in time_slots if t + duration <= 17*60])
    return c

def crossover(parent1, parent2):
    # Single-point crossover
    point = random.randint(1, len(parent1)-1)
    child = parent1[:point] + parent2[point:]
    return child

def genetic_algorithm(class_data, generations=100, pop_size=30):
    population = [build_chromosome(class_data) for _ in range(pop_size)]
    for gen in range(generations):
        population = sorted(population, key=lambda chrom: fitness(chrom, class_data), reverse=True)
        next_gen = population[:4]
        while len(next_gen) < pop_size:
            if random.random() < 0.7:
                p1, p2 = random.sample(population[:10], 2)
                child = crossover(p1, p2)
            else:
                p = random.choice(population[:10])
                child = mutate(p)
            next_gen.append(child)
        population = next_gen
        if gen % 10 == 0:
            print(f"Generation {gen}, best fitness: {fitness(population[0], class_data)}")
    return population[0]

def print_schedule(chromosome):
    print("\nBest Schedule:")
    for a in sorted(chromosome, key=lambda x: (x['assignment']['day'], x['assignment']['start_time'])):
        st = a['assignment']['start_time']
        et = st + a['duration']
        print(f"{a['assignment']['day']} {st//60:02d}:{st%60:02d}-{et//60:02d}:{et%60:02d} | {a['coursename']} | Section {a['section']} | Room {a['roomid']} | Emp {a['employeeid']}")

def chromosome_to_schedule(chromosome):
    schedule = []
    for a in chromosome:
        schedule.append({
            'course_id': a['courseid'],
            'coursename': a['coursename'],
            'section': a['section'],
            'start_time': a['assignment']['start_time'],
            'end_time': a['assignment']['start_time'] + a['duration'],
            'duration': a['duration'],
            'roomid': a['roomid'],
            'employeeid': a['employeeid'],
            'day': a['assignment']['day'],
            'start_time_str': f"{a['assignment']['start_time']//60:02d}:{a['assignment']['start_time']%60:02d}",
            'end_time_str': f"{(a['assignment']['start_time']+a['duration'])//60:02d}:{(a['assignment']['start_time']+a['duration'])%60:02d}"
        })
    return schedule

def export_schedule_excel(filename, schedule):
    wb = Workbook()
    ws = wb.active
    ws.title = "Schedule"
    header_fill = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    days = ['Time', 'Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday']
    day_to_col = {day: idx+2 for idx, day in enumerate(days[1:])}
    for col, day in enumerate(days, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = day
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
    time_slots = []
    time_to_row = {}
    current_row = 2
    for hour in range(8, 17):
        for minute in [0, 30]:
            time_str = f"{hour:02d}:{minute:02d}"
            time_to_row[time_str] = current_row
            next_minute = minute + 30
            next_hour = hour
            if next_minute >= 60:
                next_minute = 0
                next_hour += 1
            if hour < 12:
                current_time = f"{hour}:{minute:02d} AM"
            elif hour == 12:
                current_time = f"12:{minute:02d} PM"
            else:
                current_time = f"{hour-12}:{minute:02d} PM"
            if next_hour < 12:
                next_time = f"{next_hour}:{next_minute:02d} AM"
            elif next_hour == 12:
                next_time = f"12:{next_minute:02d} PM"
            else:
                next_time = f"{next_hour-12}:{next_minute:02d} PM"
            time_slots.append(f"{current_time} - {next_time}")
            current_row += 1
    for idx, time_slot in enumerate(time_slots):
        cell = ws.cell(row=idx+2, column=1)
        cell.value = time_slot
        cell.alignment = center_align
        cell.border = thin_border
    for class_info in schedule:
        start_time = class_info['start_time']
        if isinstance(start_time, int):
            hours = start_time // 60
            minutes = start_time % 60
            start_time = f"{hours:02d}:{minutes:02d}"
        start_row = time_to_row.get(start_time)
        if not start_row:
            continue
        duration = class_info['duration']
        num_slots = max(1, duration // 30)
        end_row = start_row + num_slots - 1
        day = class_info.get('day', 'Monday')
        col = day_to_col.get(day, 2)
        roomid_str = ', '.join(str(r) for r in (class_info['roomid'] if isinstance(class_info['roomid'], list) else [class_info['roomid']]))
        class_text = (
            f"{class_info['coursename']}\n"
            f"{class_info['section']}\n"
            f"Room {roomid_str}"
        )
        cell = ws.cell(row=start_row, column=col)
        if not isinstance(cell, MergedCell):
            cell.value = class_text
        cell.alignment = center_align
        cell.border = thin_border
        if num_slots > 1:
            ws.merge_cells(start_row=start_row, start_column=col, end_row=end_row, end_column=col)
            for row in range(start_row + 1, end_row + 1):
                cell = ws.cell(row=row, column=col)
                cell.alignment = center_align
                cell.border = thin_border
    for col in range(1, len(days) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 25
    for row in range(1, len(time_slots) + 2):
        ws.row_dimensions[row].height = 50
    wb.save(filename)
    print(f"\nSchedule exported to Excel file: {filename}")

def log_unscheduled_classes(class_data, schedule, log_filename='logs.txt'):
    with open(log_filename, 'w') as f:
        f.write("SCHEDULED CLASSES (EXACT MATCHES):\n")
        for s in schedule:
            f.write(f"Scheduled: {s['coursename']} | Section {s['section']} | Room {s['roomid']} | Emp {s['employeeid']} | Day {s['day']} | Duration {s['duration']}\n")
        f.write("\nREQUESTED VS ACTUAL:\n")
        for course in class_data['Courses']:
            for sched in course['classschedule']:
                sched_day = sched.get('day', '')
                sched_day_val = sched_day if isinstance(sched_day, str) else (sched_day[0] if isinstance(sched_day, list) and sched_day else '')
                sched_duration = parse_duration(sched['duration'])
                found = False
                for s in schedule:
                    if (
                        s['coursename'] == course.get('coursename', '') and
                        s['section'] == course['section'] and
                        str(s['roomid']) == str(sched['roomid']) and
                        s['employeeid'] == sched['employeeid'] and
                        s['duration'] == sched_duration and
                        (s['day'] == sched_day_val or s['day'] == (','.join(sched['day']) if isinstance(sched.get('day', ''), list) else sched.get('day', '')))
                    ):
                        found = True
                        f.write(f"REQUESTED: {course.get('coursename', '')} | Section {course['section']} | Room {sched['roomid']} | Emp {sched['employeeid']} | Day {sched.get('day', '')} | Duration {sched['duration']}\n")
                        f.write(f"  -> SCHEDULED: {s['coursename']} | Section {s['section']} | Room {s['roomid']} | Emp {s['employeeid']} | Day {s['day']} | Duration {s['duration']}\n")
                        break
                if not found:
                    f.write(f"REQUESTED: {course.get('coursename', '')} | Section {course['section']} | Room {sched['roomid']} | Emp {sched['employeeid']} | Day {sched.get('day', '')} | Duration {sched['duration']}\n")
                    f.write(f"  -> UNSCHEDULED\n")

# Example usage (replace with your data loading)
if __name__ == "__main__":
    with open('schedule.json') as f:
        class_data = json.load(f)
    if isinstance(class_data, list):
        # Flatten all Courses from all dicts in the list
        all_courses = []
        for entry in class_data:
            all_courses.extend(entry['Courses'])
        class_data = {'Courses': all_courses}
    best = genetic_algorithm(class_data, generations=100, pop_size=30)
    print_schedule(best)
    schedule = chromosome_to_schedule(best)

    # Ensure output directory exists
    output_dir = 'excel_schedules'
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    export_schedule_excel(os.path.join(output_dir, 'genetic_best_schedule.xlsx'), schedule)

    # Export per section
    sections = set(c['section'] for c in schedule)
    for section in sections:
        filtered = [c for c in schedule if c['section'] == section]
        export_schedule_excel(os.path.join(output_dir, f'section_{section}_genetic_schedule.xlsx'), filtered)

    # Export per employee
    employees = set(c['employeeid'] for c in schedule)
    for emp in employees:
        filtered = [c for c in schedule if c['employeeid'] == emp]
        export_schedule_excel(os.path.join(output_dir, f'employee_{emp}_genetic_schedule.xlsx'), filtered)

    # Export per room
    rooms = set()
    for c in schedule:
        if isinstance(c['roomid'], list):
            rooms.update(c['roomid'])
        else:
            rooms.add(c['roomid'])
    for room in rooms:
        filtered = [c for c in schedule if (room in c['roomid'] if isinstance(c['roomid'], list) else c['roomid'] == room)]
        export_schedule_excel(os.path.join(output_dir, f'room_{room}_genetic_schedule.xlsx'), filtered)

    log_unscheduled_classes(class_data, schedule) 