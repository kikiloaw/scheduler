import random
import copy
from typing import List, Dict
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os
import json
from openpyxl.cell.cell import MergedCell
from collections import defaultdict

# Helper: parse duration string to minutes
def parse_duration(duration_str):
    if ':' in duration_str:
        hours, minutes = map(int, duration_str.split(':'))
        return hours * 60 + minutes
    else:
        return int(duration_str) * 60

days_of_week = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]
time_slots = [8*60 + 30*i for i in range(18)]  # 8:00 to 17:30, 30-min slots

def random_assignment(class_entry):
    duration = parse_duration(class_entry['duration'])
    possible_days = class_entry.get('day', days_of_week)
    if isinstance(possible_days, str):
        possible_days = [possible_days]
    day = random.choice(possible_days)
    sched_type = class_entry.get('Type', 'regular').lower()
    if sched_type == 'overload':
        evening_start = 17 * 60 + 30  # 17:30
        evening_end = 20 * 60 + 30    # 20:30
        valid_start_times = [t for t in range(evening_start, evening_end + 1, 30) if t + duration <= evening_end]
    else:
        break_start = 12 * 60  # 12:00 PM
        break_end = 13 * 60    # 1:00 PM
        valid_start_times = []
        for t in time_slots:
            if t + duration <= break_start or t >= break_end:
                if t + duration <= 17*60:
                    valid_start_times.append(t)
    if not valid_start_times:
        start_time = min(time_slots)  # fallback
    else:
        half = max(1, len(valid_start_times)//2)
        start_time = random.choice(valid_start_times[:half])
    return {'day': day, 'start_time': start_time}

def build_chromosome(class_data):
    assignments = []
    for course in class_data['Courses']:
        for sched in course['classschedule']:
            assignment = random_assignment(sched)
            room_id_val = sched['roomid']
            possible_rooms = []
            chosen_room = None
            if isinstance(room_id_val, list):
                possible_rooms = room_id_val
                if possible_rooms:
                    chosen_room = random.choice(possible_rooms)
                else:
                    chosen_room = -1 
            else:
                possible_rooms = [room_id_val]
                chosen_room = room_id_val
            # Store allowed_days for use in mutation and fitness
            allowed_days = sched.get('day', days_of_week)
            if isinstance(allowed_days, str):
                allowed_days = [allowed_days]
            assignments.append({
                'courseid': course['courseid'],
                'coursename': course.get('coursename', ''),
                'section': course['section'],
                'roomid': chosen_room,
                'possible_rooms': possible_rooms,
                'employeeid': sched['employeeid'],
                'duration': parse_duration(sched['duration']),
                'assignment': assignment,
                'Type': sched.get('Type', 'regular'),
                'allowed_days': allowed_days
            })
    return assignments

def fitness(chromosome, class_data=None):
    score = 0
    scheduled = 0
    conflicts = 0
    used_days = set()
    course_day = set()
    used_time_slots = set()
    am_used = 0
    # Track assignments per section and day for hard constraint
    section_day_slots = defaultdict(lambda: defaultdict(list))
    for a in chromosome:
        section_day_slots[a['section']][a['assignment']['day']].append(a['assignment']['start_time'])
    for i, a in enumerate(chromosome):
        scheduled += 1
        used_days.add((a['section'], a['assignment']['day']))
        key = (a['section'], a['courseid'], a['assignment']['day'])
        used_time_slots.add((a['assignment']['day'], a['assignment']['start_time']))
        # Count AM slots (before 12:00 PM)
        if a['assignment']['start_time'] < 12*60:
            am_used += 1
        # Penalize if scheduled on a day not in allowed_days
        if a['assignment']['day'] not in a.get('allowed_days', days_of_week):
            score -= 20
        if key in course_day:
            score -= 10
        else:
            course_day.add(key)
        for j, b in enumerate(chromosome):
            if i == j:
                continue
            if a['section'] == b['section'] and a['assignment']['day'] == b['assignment']['day']:
                if is_time_conflict(a, b):
                    score -= 30
                    conflicts += 1
            if a['roomid'] == b['roomid'] and a['assignment']['day'] == b['assignment']['day']:
                if is_time_conflict(a, b):
                    score -= 30
                    conflicts += 1
            if a['employeeid'] == b['employeeid'] and a['assignment']['day'] == b['assignment']['day']:
                if is_time_conflict(a, b):
                    score -= 30
                    conflicts += 1
    score += scheduled
    score += 0.1 * len(used_days)
    # Reward for using more unique time slots
    score += 0.2 * len(used_time_slots)
    # Reward for using AM slots
    score += 0.1 * am_used
    # Penalize gaps for each section
    section_times = defaultdict(list)
    for a in chromosome:
        section_times[a['section']].append(a['assignment']['start_time'])
    for times in section_times.values():
        times.sort()
        for i in range(1, len(times)):
            gap = times[i] - (times[i-1] + 30)  # 30 is the slot size
            if gap > 0:
                score -= 0.05 * gap  # Penalize larger gaps more
    # Penalize gaps at the start of the day for each section
    for section, times in section_times.items():
        if times:
            earliest = min(times)
            if earliest > min(time_slots):
                score -= 0.2 * (earliest - min(time_slots))  # Penalize late start
    # Penalize using more unique days per section (stronger penalty)
    section_days = defaultdict(set)
    for a in chromosome:
        section_days[a['section']].add(a['assignment']['day'])
    # Hard constraint: if a section uses more than one day and any of those days has vacant slots, apply a very large penalty
    for section, days in section_days.items():
        if len(days) > 1:
            # For each used day, count number of slots used
            for d in days:
                slots_used = len(section_day_slots[section][d])
                if slots_used < len(time_slots):
                    score -= 10000  # Very large penalty for not filling a used day before using a new one
        score -= 15 * (len(days) - 1)  # Much stronger penalty for using more days for a section
    # Reward for using allowed days from the sample data
    for a in chromosome:
        if a['assignment']['day'] in a.get('allowed_days', days_of_week):
            score += 5  # Encourage use of allowed days
    if class_data is not None:
        input_keys = set(
            (
                course.get('coursename', ''),
                course['section'],
                str(sched['roomid']),
                sched['employeeid'],
                parse_duration(sched['duration']),
                ','.join(sched['day']) if isinstance(sched.get('day', ''), list) else sched.get('day', '')
            )
            for course in class_data['Courses'] for sched in course['classschedule']
        )
        scheduled_keys = set(
            (
                s['coursename'],
                s['section'],
                str(s['roomid']),
                s['employeeid'],
                s['duration'],
                s['day'] if isinstance(s['day'], str) else ','.join(s['day'])
            )
            for s in chromosome_to_schedule(chromosome)
        )
        missing = input_keys - scheduled_keys
        score -= 200 * len(missing)
    return score

def is_time_conflict(a, b):
    s1, d1 = a['assignment']['start_time'], a['duration']
    s2, d2 = b['assignment']['start_time'], b['duration']
    e1, e2 = s1 + d1, s2 + d2
    return not (e1 <= s2 or e2 <= s1)

def mutate(chromosome):
    c = copy.deepcopy(chromosome)
    idx = random.randint(0, len(c) - 1)
    mut_options = ['day', 'time']
    if len(c[idx].get('possible_rooms', [])) > 1:
        mut_options.append('room')
    mutation_choice = random.choice(mut_options)
    if mutation_choice == 'day':
        # Only mutate to allowed days for this class
        allowed_days = c[idx].get('allowed_days', days_of_week)
        if isinstance(allowed_days, str):
            allowed_days = [allowed_days]
        c[idx]['assignment']['day'] = random.choice(allowed_days)
    elif mutation_choice == 'room':
        possible_rooms = c[idx]['possible_rooms']
        current_room = c[idx]['roomid']
        other_rooms = [r for r in possible_rooms if r != current_room]
        if other_rooms:
            c[idx]['roomid'] = random.choice(other_rooms)
    else:  # 'time'
        duration = c[idx]['duration']
        sched_type = c[idx].get('Type', 'regular').lower()
        if sched_type == 'overload':
            evening_start = 17 * 60 + 30
            evening_end = 20 * 60 + 30
            valid_start_times = [t for t in range(evening_start, evening_end + 1, 30) if t + duration <= evening_end]
        else:
            break_start = 12 * 60
            break_end = 13 * 60
            valid_start_times = []
            for t in time_slots:
                if t + duration <= break_start or t >= break_end:
                    if t + duration <= 17 * 60:
                        valid_start_times.append(t)
        if valid_start_times:
            c[idx]['assignment']['start_time'] = random.choice(valid_start_times)
        else:
            c[idx]['assignment']['start_time'] = min(time_slots)
    return c

def crossover(parent1, parent2):
    # Single-point crossover
    point = random.randint(1, len(parent1)-1)
    child = parent1[:point] + parent2[point:]
    return child

def genetic_algorithm(class_data, generations=100, pop_size=30):
    population = [build_chromosome(class_data) for _ in range(pop_size)]
    for gen in range(generations):
        population = sorted(population, key=lambda chrom: fitness(chrom, class_data), reverse=True)
        next_gen = population[:4]
        while len(next_gen) < pop_size:
            if random.random() < 0.7:
                p1, p2 = random.sample(population[:10], 2)
                child = crossover(p1, p2)
            else:
                p = random.choice(population[:10])
                child = mutate(p)
            next_gen.append(child)
        population = next_gen
        if gen % 10 == 0:
            print(f"Generation {gen}, best fitness: {fitness(population[0], class_data)}")
    return population[0]

def print_schedule(chromosome):
    print("\nBest Schedule:")
    for a in sorted(chromosome, key=lambda x: (x['assignment']['day'], x['assignment']['start_time'])):
        st = a['assignment']['start_time']
        et = st + a['duration']
        print(f"{a['assignment']['day']} {st//60:02d}:{st%60:02d}-{et//60:02d}:{et%60:02d} | {a['coursename']} | Section {a['section']} | Room {a['roomid']} | Emp {a['employeeid']}")

def chromosome_to_schedule(chromosome):
    schedule = []
    for a in chromosome:
        schedule.append({
            'course_id': a['courseid'],
            'coursename': a['coursename'],
            'section': a['section'],
            'start_time': a['assignment']['start_time'],
            'end_time': a['assignment']['start_time'] + a['duration'],
            'duration': a['duration'],
            'roomid': a['roomid'],
            'employeeid': a['employeeid'],
            'day': a['assignment']['day'],
            'start_time_str': f"{a['assignment']['start_time']//60:02d}:{a['assignment']['start_time']%60:02d}",
            'end_time_str': f"{(a['assignment']['start_time']+a['duration'])//60:02d}:{(a['assignment']['start_time']+a['duration'])%60:02d}"
        })
    return schedule

def export_schedule_excel(filename, schedule):
    wb = Workbook()
    ws = wb.active
    ws.title = "Schedule"
    header_fill = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    days = ['Time', 'Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
    day_to_col = {day: idx+2 for idx, day in enumerate(days[1:])}
    for col, day in enumerate(days, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = day
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
    time_slots = []
    time_to_row = {}
    current_row = 2
    for hour in range(8, 21):  # 8:00 AM to 8:30 PM
        for minute in [0, 30]:
            time_str = f"{hour:02d}:{minute:02d}"
            time_to_row[time_str] = current_row
            next_minute = minute + 30
            next_hour = hour
            if next_minute >= 60:
                next_minute = 0
                next_hour += 1
            if hour < 12:
                current_time = f"{hour}:{minute:02d} AM"
            elif hour == 12:
                current_time = f"12:{minute:02d} PM"
            else:
                current_time = f"{hour-12}:{minute:02d} PM"
            if next_hour < 12:
                next_time = f"{next_hour}:{next_minute:02d} AM"
            elif next_hour == 12:
                next_time = f"12:{next_minute:02d} PM"
            else:
                next_time = f"{next_hour-12}:{next_minute:02d} PM"
            time_slots.append(f"{current_time} - {next_time}")
            current_row += 1
    for idx, time_slot in enumerate(time_slots):
        cell = ws.cell(row=idx+2, column=1)
        cell.value = time_slot
        cell.alignment = center_align
        cell.border = thin_border
    for class_info in schedule:
        start_time = class_info['start_time']
        if isinstance(start_time, int):
            hours = start_time // 60
            minutes = start_time % 60
            start_time = f"{hours:02d}:{minutes:02d}"
        start_row = time_to_row.get(start_time)
        if not start_row:
            continue
        duration = class_info['duration']
        num_slots = max(1, duration // 30)
        end_row = start_row + num_slots - 1
        day = class_info.get('day', 'Monday')
        col = day_to_col.get(day, 2)
        roomid_str = ', '.join(str(r) for r in (class_info['roomid'] if isinstance(class_info['roomid'], list) else [class_info['roomid']]))
        class_text = (
            f"{class_info['coursename']}\n"
            f"{class_info['section']}\n"
            f"Room {roomid_str}"
        )
        cell = ws.cell(row=start_row, column=col)
        if not isinstance(cell, MergedCell):
            cell.value = class_text
        cell.alignment = center_align
        cell.border = thin_border
        if num_slots > 1:
            ws.merge_cells(start_row=start_row, start_column=col, end_row=end_row, end_column=col)
            for row in range(start_row + 1, end_row + 1):
                cell = ws.cell(row=row, column=col)
                cell.alignment = center_align
                cell.border = thin_border
    for col in range(1, len(days) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 25
    for row in range(1, len(time_slots) + 2):
        ws.row_dimensions[row].height = 50
    wb.save(filename)
    print(f"\nSchedule exported to Excel file: {filename}")

def log_unscheduled_classes(class_data, schedule, log_filename='logs.txt'):
    with open(log_filename, 'w') as f:
        for course in class_data['Courses']:
            for sched in course['classschedule']:
                sched_day = sched.get('day', '')
                sched_day_val = sched_day if isinstance(sched_day, str) else (sched_day[0] if isinstance(sched_day, list) and sched_day else '')
                sched_duration = parse_duration(sched['duration'])
                found = False
                for s in schedule:
                    room_match = False
                    input_roomid = sched['roomid']
                    scheduled_roomid = s['roomid']
                    if isinstance(input_roomid, list):
                        if scheduled_roomid in input_roomid:
                            room_match = True
                    elif scheduled_roomid == input_roomid:
                        room_match = True
                    
                    if (
                        s['coursename'] == course.get('coursename', '') and
                        s['section'] == course['section'] and
                        room_match and
                        s['employeeid'] == sched['employeeid'] and
                        s['duration'] == sched_duration and
                        (s['day'] == sched_day_val or s['day'] == (','.join(sched['day']) if isinstance(sched.get('day', ''), list) else sched.get('day', '')))
                    ):
                        found = True
                        break
                if not found:
                    f.write(f"UNSCHEDULED: {course.get('coursename', '')} | Section {course['section']} | Room {sched['roomid']} | Emp {sched['employeeid']} | Day {sched.get('day', '')} | Duration {sched['duration']}\n")

# Example usage (replace with your data loading)
if __name__ == "__main__":
    with open('bsis4th year.json') as f:
        class_data = json.load(f)
    if isinstance(class_data, list):
        # Flatten all Courses from all dicts in the list
        all_courses = []
        for entry in class_data:
            all_courses.extend(entry['Courses'])
        class_data = {'Courses': all_courses}
    best = genetic_algorithm(class_data, generations=300, pop_size=100)
    print_schedule(best)
    schedule = chromosome_to_schedule(best)

    # Ensure output directory exists
    output_dir = 'excel_schedules'
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    export_schedule_excel(os.path.join(output_dir, 'genetic_best_schedule.xlsx'), schedule)

    # Export per section
    sections = set(c['section'] for c in schedule)
    for section in sections:
        filtered = [c for c in schedule if c['section'] == section]
        export_schedule_excel(os.path.join(output_dir, f'section_{section}_genetic_schedule.xlsx'), filtered)

    # Export per employee
    employees = set(c['employeeid'] for c in schedule)
    for emp in employees:
        filtered = [c for c in schedule if c['employeeid'] == emp]
        export_schedule_excel(os.path.join(output_dir, f'employee_{emp}_genetic_schedule.xlsx'), filtered)

    # Export per room
    rooms = set()
    for c in schedule:
        if isinstance(c['roomid'], list):
            rooms.update(c['roomid'])
        else:
            rooms.add(c['roomid'])
    for room in rooms:
        filtered = [c for c in schedule if (room in c['roomid'] if isinstance(c['roomid'], list) else c['roomid'] == room)]
        export_schedule_excel(os.path.join(output_dir, f'room_{room}_genetic_schedule.xlsx'), filtered)

    log_unscheduled_classes(class_data, schedule) 