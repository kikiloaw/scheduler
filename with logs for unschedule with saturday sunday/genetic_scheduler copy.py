import random
import copy
from typing import List, Dict
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os
import json
from openpyxl.cell.cell import MergedCell

# Scheduler settings for allowing/disallowing conflicts
SCHEDULER_SETTINGS = {
    "allow_room_conflict": True,      # If True, allows two classes in the same room at the same time
    "allow_employee_conflict": False, # If True, allows an employee to teach two classes at the same time
    # Add more settings as needed
}

# Helper: parse duration string to minutes
def parse_duration(duration_str):
    if ':' in duration_str:
        hours, minutes = map(int, duration_str.split(':'))
        return hours * 60 + minutes
    else:
        return int(duration_str) * 60

days_of_week = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]
time_slots = [8*60 + 30*i for i in range(18)]  # 8:00 to 17:30, 30-min slots

def has_conflict(start_time, duration, day, roomid, employeeid, section, assignments):
    end_time = start_time + duration
    for a in assignments:
        if a['assignment']['day'] != day:
            continue
        # Room conflict
        if not SCHEDULER_SETTINGS.get("allow_room_conflict", False):
            if a['roomid'] == roomid:
                s2, d2 = a['assignment']['start_time'], a['duration']
                e2 = s2 + d2
                if not (end_time <= s2 or (start_time >= e2)):
                    return True
        # Employee conflict
        if not SCHEDULER_SETTINGS.get("allow_employee_conflict", False):
            if a['employeeid'] == employeeid:
                s2, d2 = a['assignment']['start_time'], a['duration']
                e2 = s2 + d2
                if not (end_time <= s2 or (start_time >= e2)):
                    return True
        # Section conflict (always enforced)
        if a['section'] == section:
            s2, d2 = a['assignment']['start_time'], a['duration']
            e2 = s2 + d2
            if not (end_time <= s2 or (start_time >= e2)):
                return True
    return False

def find_earliest_available_slot(duration, day, roomid, employeeid, section, assignments, sched_type):
    # Determine start and end times based on type
    if sched_type == 'overload':
        start_time = 17 * 60 + 30  # 5:30 PM
        end_time = 20 * 60 + 30    # 8:30 PM
        # Always try 5:30 PM first
        if not has_conflict(start_time, duration, day, roomid, employeeid, section, assignments):
            return start_time
        t = start_time + 30
        while t + duration <= end_time:
            if not has_conflict(t, duration, day, roomid, employeeid, section, assignments):
                return t
            t += 30
        return None
    else:
        # Morning block: always try 8:00 AM first
        morning_start = 8 * 60
        morning_end = 12 * 60  # up to 12:00 PM
        if not has_conflict(morning_start, duration, day, roomid, employeeid, section, assignments):
            return morning_start
        t = morning_start + 30
        while t + duration <= morning_end:
            if not has_conflict(t, duration, day, roomid, employeeid, section, assignments):
                return t
            t += 30
        # Afternoon block: always try 1:00 PM first
        afternoon_start = 13 * 60
        afternoon_end = 17 * 60  # up to 5:00 PM
        if not has_conflict(afternoon_start, duration, day, roomid, employeeid, section, assignments):
            return afternoon_start
        t = afternoon_start + 30
        while t + duration <= afternoon_end:
            if not has_conflict(t, duration, day, roomid, employeeid, section, assignments):
                return t
            t += 30
        return None

def find_block_aligned_slot(duration, day, roomid, employeeid, section, assignments, sched_type, day_block_sizes):
    # Determine block size for this day
    if day in day_block_sizes:
        block_size = day_block_sizes[day]
    else:
        block_size = duration
        day_block_sizes[day] = block_size
    # Determine start and end times based on type
    if sched_type == 'overload':
        start_time = 17 * 60 + 30  # 5:30 PM
        end_time = 20 * 60 + 30    # 8:30 PM
        # Always try 5:30 PM first
        if not has_conflict(start_time, duration, day, roomid, employeeid, section, assignments):
            return start_time
        t = start_time + block_size
        while t + duration <= end_time:
            if not has_conflict(t, duration, day, roomid, employeeid, section, assignments):
                return t
            t += block_size
        return None
    else:
        # Morning block: always try 8:00 AM first
        morning_start = 8 * 60
        morning_end = 12 * 60  # up to 12:00 PM
        if not has_conflict(morning_start, duration, day, roomid, employeeid, section, assignments):
            return morning_start
        t = morning_start + block_size
        while t + duration <= morning_end:
            if not has_conflict(t, duration, day, roomid, employeeid, section, assignments):
                return t
            t += block_size
        # Afternoon block: always try 1:00 PM first
        afternoon_start = 13 * 60
        afternoon_end = 17 * 60  # up to 5:00 PM
        if not has_conflict(afternoon_start, duration, day, roomid, employeeid, section, assignments):
            return afternoon_start
        t = afternoon_start + block_size
        while t + duration <= afternoon_end:
            if not has_conflict(t, duration, day, roomid, employeeid, section, assignments):
                return t
            t += block_size
        return None

def random_assignment(class_entry, assignments_so_far=None, section=None, day_block_sizes=None):
    if assignments_so_far is None:
        assignments_so_far = []
    if day_block_sizes is None:
        day_block_sizes = {}
    duration = parse_duration(class_entry['duration'])
    possible_days = class_entry.get('day', days_of_week)
    if isinstance(possible_days, str):
        possible_days = [possible_days]
    day = random.choice(possible_days)
    sched_type = class_entry.get('Type', 'regular').lower()
    roomid = class_entry['roomid'][0] if isinstance(class_entry['roomid'], list) else class_entry['roomid']
    section_val = section if section is not None else class_entry.get('section', None)
    start_time = find_block_aligned_slot(duration, day, roomid, class_entry['employeeid'], section_val, assignments_so_far, sched_type, day_block_sizes)
    if start_time is None:
        # fallback: just use 8:00 AM or 5:30 PM
        start_time = 8*60 if sched_type != 'overload' else 17*60+30
    return {'day': day, 'start_time': start_time}

def build_chromosome(class_data):
    assignments = []
    day_block_sizes = {}
    for course in class_data['Courses']:
        for sched in course['classschedule']:
            assignment = random_assignment(sched, assignments, section=course['section'], day_block_sizes=day_block_sizes)
            room_id_val = sched['roomid']
            possible_rooms = []
            chosen_room = None
            if isinstance(room_id_val, list):
                possible_rooms = room_id_val
                if possible_rooms:
                    chosen_room = possible_rooms[0]  # always pick first for tight packing
                else:
                    chosen_room = -1 
            else:
                possible_rooms = [room_id_val]
                chosen_room = room_id_val
            assignments.append({
                'ClassID': course['ClassID'],
                'coursename': course.get('coursename', ''),
                'section': course['section'],
                'roomid': chosen_room,
                'possible_rooms': possible_rooms,
                'employeeid': sched['employeeid'],
                'duration': parse_duration(sched['duration']),
                'assignment': assignment,
                'Type': sched.get('Type', 'regular'),
                'name': sched.get('name', None)
            })
    return assignments

def fitness(chromosome, class_data=None, log_conflicts=True):
    score = 0
    scheduled = 0
    conflicts = 0
    used_days = set()
    course_day = set()
    
    # Track all conflicts to summarize them
    all_conflicts = []
    
    for i, a in enumerate(chromosome):
        scheduled += 1
        used_days.add((a['section'], a['assignment']['day']))
        key = (a['section'], a['ClassID'], a['assignment']['day'])
        if key in course_day:
            score -= 10
            if log_conflicts:
                all_conflicts.append((a, None, "Same course scheduled twice on same day"))
        else:
            course_day.add(key)
            
        for j, b in enumerate(chromosome):
            if i == j:
                continue
            if a['section'] == b['section'] and a['assignment']['day'] == b['assignment']['day']:
                if is_time_conflict(a, b):
                    score -= 30
                    conflicts += 1
                    if log_conflicts:
                        all_conflicts.append((a, b, "Section time conflict"))
                        
            if a['roomid'] == b['roomid'] and a['assignment']['day'] == b['assignment']['day']:
                if is_time_conflict(a, b):
                    score -= 30
                    conflicts += 1
                    if log_conflicts:
                        all_conflicts.append((a, b, f"Room {a['roomid']} conflict"))
                        
            if a['employeeid'] == b['employeeid'] and a['assignment']['day'] == b['assignment']['day']:
                if is_time_conflict(a, b):
                    score -= 30
                    conflicts += 1
                    if log_conflicts:
                        all_conflicts.append((a, b, "Employee time conflict"))
    
    score += scheduled
    score += 0.1 * len(used_days)
    
    if class_data is not None:
        input_keys = set(
            (
                course.get('coursename', ''),
                course['section'],
                str(sched['roomid']),
                sched['employeeid'],
                parse_duration(sched['duration']),
                ','.join(sched['day']) if isinstance(sched.get('day', ''), list) else sched.get('day', '')
            )
            for course in class_data['Courses'] for sched in course['classschedule']
        )
        scheduled_keys = set(
            (
                s['coursename'],
                s['section'],
                str(s['roomid']),
                s['employeeid'],
                s['duration'],
                s['day'] if isinstance(s['day'], str) else ','.join(s['day'])
            )
            for s in chromosome_to_schedule(chromosome)
        )
        missing = input_keys - scheduled_keys
        if missing and log_conflicts:
            for m in missing:
                pass  # log_unscheduled removed; handled elsewhere
        score -= 200 * len(missing)
    
    # Log all conflicts at once if we're logging
    if log_conflicts and all_conflicts:
        with open('logs.txt', 'a') as f:
            f.write("\n=== Detailed Conflicts ===\n\n")
            for a, b, conflict_type in all_conflicts:
                if b:  # If there's a conflicting class
                    conflict_msg = f"CONFLICT: {a['coursename']} ({a['section']}) vs {b['coursename']} ({b['section']})\n"
                    conflict_msg += f"  Type: {conflict_type}\n"
                    conflict_msg += f"  Day: {a['assignment']['day']}\n"
                    conflict_msg += f"  Time: {format_time(a['assignment']['start_time'])}-{format_time(a['assignment']['start_time'] + a['duration'])} vs "
                    conflict_msg += f"{format_time(b['assignment']['start_time'])}-{format_time(b['assignment']['start_time'] + b['duration'])}\n"
                    conflict_msg += f"  Room: {a['roomid']}\n"
                    conflict_msg += f"  Employee: {a['employeeid']}\n\n"
                else:  # For same course conflicts
                    conflict_msg = f"CONFLICT: {a['coursename']} ({a['section']})\n"
                    conflict_msg += f"  Type: {conflict_type}\n"
                    conflict_msg += f"  Day: {a['assignment']['day']}\n"
                    conflict_msg += f"  Time: {format_time(a['assignment']['start_time'])}-{format_time(a['assignment']['start_time'] + a['duration'])}\n"
                    conflict_msg += f"  Room: {a['roomid']}\n"
                    conflict_msg += f"  Employee: {a['employeeid']}\n\n"
                f.write(conflict_msg)
    
    return score

def format_time(minutes):
    hours = minutes // 60
    mins = minutes % 60
    return f"{hours:02d}:{mins:02d}"

def is_time_conflict(a, b):
    s1, d1 = a['assignment']['start_time'], a['duration']
    s2, d2 = b['assignment']['start_time'], b['duration']
    e1, e2 = s1 + d1, s2 + d2
    return not (e1 <= s2 or e2 <= s1)

def mutate(chromosome):
    c = copy.deepcopy(chromosome)
    idx = random.randint(0, len(c) - 1)
    mut_options = ['day', 'time']
    if len(c[idx].get('possible_rooms', [])) > 1:
        mut_options.append('room')
    mutation_choice = random.choice(mut_options)
    if mutation_choice == 'day':
        allowed_days = c[idx].get('allowed_days', days_of_week)
        if isinstance(allowed_days, str):
            allowed_days = [allowed_days]
        c[idx]['assignment']['day'] = random.choice(allowed_days)
    elif mutation_choice == 'room':
        possible_rooms = c[idx]['possible_rooms']
        current_room = c[idx]['roomid']
        other_rooms = [r for r in possible_rooms if r != current_room]
        if other_rooms:
            c[idx]['roomid'] = random.choice(other_rooms)
    else:  # 'time'
        duration = c[idx]['duration']
        sched_type = c[idx].get('Type', 'regular').lower()
        if sched_type == 'overload':
            evening_start = 17 * 60 + 30
            evening_end = 20 * 60 + 30
            valid_start_times = [t for t in range(evening_start, evening_end + 1, 30) if t + duration <= evening_end]
        else:
            break_start = 12 * 60
            break_end = 13 * 60
            valid_start_times = []
            for t in time_slots:
                if t + duration <= break_start or t >= break_end:
                    if t + duration <= 17 * 60:
                        valid_start_times.append(t)
        if valid_start_times:
            c[idx]['assignment']['start_time'] = random.choice(valid_start_times)
        else:
            c[idx]['assignment']['start_time'] = min(time_slots)
    return c

def crossover(parent1, parent2):
    # Single-point crossover
    point = random.randint(1, len(parent1)-1)
    child = parent1[:point] + parent2[point:]
    return child

def genetic_algorithm(class_data, generations=100, pop_size=30):
    # Don't write to log file at start - we only want final results
    population = [build_chromosome(class_data) for _ in range(pop_size)]
    best_fitness = float('-inf')
    best_solution = None
    
    for gen in range(generations):
        population = sorted(population, key=lambda chrom: fitness(chrom, class_data, log_conflicts=False), reverse=True)
        current_best = population[0]
        current_fitness = fitness(current_best, class_data, log_conflicts=False)
        
        if current_fitness > best_fitness:
            best_fitness = current_fitness
            best_solution = current_best
            
        next_gen = population[:4]  # Keep the best 4
        
        while len(next_gen) < pop_size:
            if random.random() < 0.7:  # 70% chance of crossover
                p1 = random.choice(population[:10])
                p2 = random.choice(population[:10])
                child = crossover(p1, p2)
            else:
                child = mutate(random.choice(population[:10]))
            next_gen.append(child)
            
        population = next_gen
    
    # Now log only the final conflicts for the best solution
    with open('logs.txt', 'w') as f:
        f.write("=== Final Schedule Conflicts ===\n\n")
    
    # Get conflicts from the best solution
    fitness(best_solution, class_data, log_conflicts=True)
    
    return best_solution

def print_schedule(chromosome):
    print("\nBest Schedule:")
    for a in sorted(chromosome, key=lambda x: (x['assignment']['day'], x['assignment']['start_time'])):
        st = a['assignment']['start_time']
        et = st + a['duration']
        print(f"{a['assignment']['day']} {st//60:02d}:{st%60:02d}-{et//60:02d}:{et%60:02d} | {a['coursename']} | Section {a['section']} | Room {a['roomid']} | Emp {a['employeeid']}")

def chromosome_to_schedule(chromosome):
    schedule = []
    for a in chromosome:
        schedule.append({
            'course_id': a['ClassID'],
            'coursename': a['coursename'],
            'section': a['section'],
            'start_time': a['assignment']['start_time'],
            'end_time': a['assignment']['start_time'] + a['duration'],
            'duration': a['duration'],
            'roomid': a['roomid'],
            'employeeid': a['employeeid'],
            'day': a['assignment']['day'],
            'start_time_str': f"{a['assignment']['start_time']//60:02d}:{a['assignment']['start_time']%60:02d}",
            'end_time_str': f"{(a['assignment']['start_time']+a['duration'])//60:02d}:{(a['assignment']['start_time']+a['duration'])%60:02d}"
        })
    return schedule

def export_schedule_excel(filename, schedule, plotted_set=None):
    wb = Workbook()
    ws = wb.active
    ws.title = "Schedule"
    header_fill = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    days = ['Time', 'Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
    day_to_col = {day: idx+2 for idx, day in enumerate(days[1:])}
    for col, day in enumerate(days, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = day
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
    time_slots = []
    time_to_row = {}
    current_row = 2
    for hour in range(8, 21):  # 8:00 AM to 8:30 PM
        for minute in [0, 30]:
            time_str = f"{hour:02d}:{minute:02d}"
            time_to_row[time_str] = current_row
            next_minute = minute + 30
            next_hour = hour
            if next_minute >= 60:
                next_minute = 0
                next_hour += 1
            if hour < 12:
                current_time = f"{hour}:{minute:02d} AM"
            elif hour == 12:
                current_time = f"12:{minute:02d} PM"
            else:
                current_time = f"{hour-12}:{minute:02d} PM"
            if next_hour < 12:
                next_time = f"{next_hour}:{next_minute:02d} AM"
            elif next_hour == 12:
                next_time = f"12:{next_minute:02d} PM"
            else:
                next_time = f"{next_hour-12}:{next_minute:02d} PM"
            time_slots.append(f"{current_time} - {next_time}")
            current_row += 1
    for idx, time_slot in enumerate(time_slots):
        cell = ws.cell(row=idx+2, column=1)
        cell.value = time_slot
        cell.alignment = center_align
        cell.border = thin_border
    if plotted_set is not None:
        # Track plotted classes
        def make_key(class_info):
            return (
                class_info['coursename'],
                class_info['section'],
                class_info['employeeid'],
                class_info['duration'],
                class_info['day'],
                str(class_info['roomid'])
            )
    # Define a color palette for courses
    color_palette = [
        'FFB6C1', 'ADD8E6', '90EE90', 'FFD700', 'FFA07A', '20B2AA', '9370DB', 'F08080', 'E0FFFF', 'FFE4E1',
        'B0E0E6', 'FF6347', '4682B4', 'D2B48C', '9ACD32', '40E0D0', 'FF69B4', 'CD5C5C', '00CED1', '1E90FF',
        'B22222', 'FF7F50', '6A5ACD', '00FA9A', '7B68EE', '00FF7F', 'DC143C', '00BFFF', 'FF8C00', '8A2BE2'
    ]
    course_colors = {}
    color_index = 0

    for class_info in schedule:
        start_time = class_info['start_time']
        if isinstance(start_time, int):
            hours = start_time // 60
            minutes = start_time % 60
            start_time = f"{hours:02d}:{minutes:02d}"
        start_row = time_to_row.get(start_time)
        if not start_row:
            print(f"[DEBUG] Skipping class (no start_row): {class_info}")
            continue
        duration = class_info['duration']
        # Validate duration is a multiple of 30
        if duration % 30 != 0:
            print(f"[WARNING] Class {class_info['coursename']} {class_info['section']} has non-30-min-multiple duration: {duration}")
            continue
        num_slots = max(1, duration // 30)
        end_row = start_row + num_slots - 1
        day = class_info.get('day', 'Monday')
        col = day_to_col.get(day, 2)
        roomid_str = ', '.join(str(r) for r in (class_info['roomid'] if isinstance(class_info['roomid'], list) else [class_info['roomid']]))
        emp_display = class_info.get('name') or f"Emp: {class_info['employeeid']}"
        class_text = (
            f"{class_info['coursename']}\n"
            f"{class_info['section']}\n"
            f"Room {roomid_str}\n"
            f"{emp_display}"
        )
        print(f"[DEBUG] Class: {class_info['coursename']} | Section: {class_info['section']} | Day: {day} | Start: {start_time} | Duration: {duration} | StartRow: {start_row} | EndRow: {end_row} | NumSlots: {num_slots} | Col: {col}")
        # Check for overlap before merging
        overlap = False
        for row in range(start_row, end_row + 1):
            cell = ws.cell(row=row, column=col)
            if cell.value not in (None, ""):
                print(f"[WARNING] Overlap detected for {class_info['coursename']} {class_info['section']} at row {row}, col {col}")
                overlap = True
                break
        if overlap:
            continue  # Skip merging this class
        # Assign a color for this course if not already assigned
        course_name = class_info['coursename']
        if course_name not in course_colors:
            course_colors[course_name] = color_palette[color_index % len(color_palette)]
            color_index += 1
        fill = PatternFill(start_color=course_colors[course_name], end_color=course_colors[course_name], fill_type='solid')
        cell = ws.cell(row=start_row, column=col)
        if not isinstance(cell, MergedCell):
            cell.value = class_text
        cell.fill = fill
        cell.alignment = center_align
        cell.border = thin_border
        if num_slots > 1:
            print(f"[DEBUG] Merging rows {start_row} to {end_row} in column {col} for class {class_info['coursename']} {class_info['section']}")
            ws.merge_cells(start_row=start_row, start_column=col, end_row=end_row, end_column=col)
            for row in range(start_row + 1, end_row + 1):
                cell = ws.cell(row=row, column=col)
                cell.fill = fill
                cell.alignment = center_align
                cell.border = thin_border
        if plotted_set is not None:
            plotted_set.add((
                class_info['coursename'],
                class_info['section'],
                class_info['employeeid'],
                class_info['duration'],
                class_info['day'],
                str(class_info['roomid'])
            ))
    for col in range(1, len(days) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 25
    for row in range(1, len(time_slots) + 2):
        ws.row_dimensions[row].height = 50
    wb.save(filename)
    print(f"\nSchedule exported to Excel file: {filename}")

def log_unplotted_classes(class_data, plotted_set, log_filename='logs.txt'):
    with open(log_filename, 'w') as f:
        f.write("=== UNSCHEDULED (UNPLOTTED) CLASSES ===\n\n")
    
    # Build a schedule from plotted classes for conflict analysis
    plotted_schedule = []
    for key in plotted_set:
        coursename, section, employeeid, duration, day, roomid = key
        plotted_schedule.append({
            'coursename': coursename,
            'section': section,
            'employeeid': employeeid,
            'duration': duration,
            'day': day,
            'roomid': roomid,
            'start_time': 0  # We don't have exact times, but we can check day conflicts
        })
    
    for course in class_data['Courses']:
        for sched in course['classschedule']:
            allowed_days = sched.get('day', days_of_week)
            if isinstance(allowed_days, str):
                allowed_days = [allowed_days]
            sched_duration = parse_duration(sched['duration'])
            allowed_rooms = sched['roomid'] if isinstance(sched['roomid'], list) else [sched['roomid']]
            found = False
            for day in allowed_days:
                for room in allowed_rooms:
                    key = (
                        course.get('coursename', ''),
                        course['section'],
                        sched['employeeid'],
                        sched_duration,
                        day,
                        str(room)
                    )
                    if key in plotted_set:
                        found = True
                        break
                if found:
                    break
            
            if not found:
                # Analyze conflicts for this unscheduled class
                conflicts = []
                
                # Check room conflicts
                for day in allowed_days:
                    for room in allowed_rooms:
                        room_conflicts = [c for c in plotted_schedule 
                                        if c['day'] == day and str(c['roomid']) == str(room)]
                        if room_conflicts and not SCHEDULER_SETTINGS.get('allow_room_conflict', False):
                            conflicts.append(f"Room {room} conflict on {day} with: " + 
                                          ", ".join([f"{c['coursename']} {c['section']}" for c in room_conflicts]))
                
                # Check employee conflicts
                for day in allowed_days:
                    emp_conflicts = [c for c in plotted_schedule 
                                   if c['day'] == day and c['employeeid'] == sched['employeeid']]
                    if emp_conflicts and not SCHEDULER_SETTINGS.get('allow_employee_conflict', False):
                        conflicts.append(f"Employee {sched['employeeid']} conflict on {day} with: " + 
                                      ", ".join([f"{c['coursename']} {c['section']}" for c in emp_conflicts]))
                
                # Check section conflicts (same course/section on same day)
                for day in allowed_days:
                    section_conflicts = [c for c in plotted_schedule 
                                       if c['day'] == day and c['section'] == course['section'] 
                                       and c['coursename'] == course.get('coursename', '')]
                    if section_conflicts:
                        conflicts.append(f"Section {course['section']} conflict on {day} with: " + 
                                      ", ".join([f"{c['coursename']} {c['section']}" for c in section_conflicts]))
                
                with open(log_filename, 'a') as f:
                    f.write(f"\nUNSCHEDULED CLASS:\n")
                    f.write(f"Course: {course.get('coursename', '')}\n")
                    f.write(f"Section: {course['section']}\n")
                    f.write(f"Room: {sched['roomid']}\n")
                    f.write(f"Employee: {sched['employeeid']}\n")
                    f.write(f"Day: {sched.get('day', '')}\n")
                    f.write(f"Duration: {sched['duration']}\n")
                    if conflicts:
                        f.write(f"CONFLICTS:\n")
                        for conflict in conflicts:
                            f.write(f"  - {conflict}\n")
                    else:
                        f.write(f"REASON: No available time slots found\n")
                    f.write("\n" + "="*50 + "\n")

# Example usage (replace with your data loading)
if __name__ == "__main__":
    with open('schedule.json') as f:
        class_data = json.load(f)
    if isinstance(class_data, list):
        # Flatten all Courses from all dicts in the list
        all_courses = []
        for entry in class_data:
            all_courses.extend(entry['Courses'])
        class_data = {'Courses': all_courses}
    best = genetic_algorithm(class_data, generations=300, pop_size=100)
    print_schedule(best)
    schedule = chromosome_to_schedule(best)

    # Track plotted classes
    plotted_set = set()
    export_schedule_excel(os.path.join('excel_schedules', 'genetic_best_schedule.xlsx'), schedule, plotted_set=plotted_set)

    # Export per section
    sections = set(c['section'] for c in schedule)
    for section in sections:
        filtered = [c for c in schedule if c['section'] == section]
        export_schedule_excel(os.path.join('excel_schedules', f'section_{section}_genetic_schedule.xlsx'), filtered, plotted_set=plotted_set)

    # Export per employee
    employees = set(c['employeeid'] for c in schedule)
    for emp in employees:
        filtered = [c for c in schedule if c['employeeid'] == emp]
        export_schedule_excel(os.path.join('excel_schedules', f'employee_{emp}_genetic_schedule.xlsx'), filtered, plotted_set=plotted_set)

    # Export per room
    rooms = set()
    for c in schedule:
        if isinstance(c['roomid'], list):
            rooms.update(c['roomid'])
        else:
            rooms.add(c['roomid'])
    for room in rooms:
        filtered = [c for c in schedule if (room in c['roomid'] if isinstance(c['roomid'], list) else c['roomid'] == room)]
        export_schedule_excel(os.path.join('excel_schedules', f'room_{room}_genetic_schedule.xlsx'), filtered, plotted_set=plotted_set)

    # Log only unplotted classes
    log_unplotted_classes(class_data, plotted_set) 